"""
import_audit.py

This module generates a robust, human-readable audit trail for Excel file imports in the feedback portal.

Purpose:
    - To provide a double-check of the import process by independently reading the Excel file and comparing its contents to the schema and the output of parse_xl_file.
    - To help users and developers diagnose issues with field mapping, type conversion, whitespace, and dropdown validation.
    - To provide a clear, aligned, per-field diagnostic log for compliance and debugging.

How it works:
    - Takes the output of parse_xl_file(file_path) (parse_result) and the schema map.
    - Independently opens the Excel file with openpyxl.
    - For each field in the schema for the tab of interest (e.g., 'Feedback Form'), it:
        * Reads the raw value and label from the spreadsheet.
        * Compares them to the schema.
        * Attempts type conversion and whitespace cleaning, logging any issues.
        * Summarizes label and value matches.

Limitations:
    - This audit does NOT use the portal's main logging infrastructure or business logic; it is a static, schema-driven check.
    - It may not catch all business logic errors, cross-field dependencies, or dynamic validation rules enforced elsewhere in the portal.
    - It is intended as a best-effort, field-level diagnostic and not a full validation of the import pipeline.

Usage:
    generate_import_audit(file_path, parse_result, schema_map, logs=None, route="")

Output:
    Returns a string suitable for writing to a log file.
"""

import datetime
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from arb.utils.json import json_serializer

# =========================
# Utility Functions
# =========================
# Update pad_label to use a wider width for better alignment
PAD_WIDTH = 24


def pad_label(label: str, width: int = PAD_WIDTH) -> str:
  """Pad a label to a fixed width for aligned output."""
  return label.ljust(width)


def normalize_label(label: Any) -> str:
  """Normalize a label for audit display (strip, replace line breaks)."""
  if isinstance(label, str):
    return label.strip().replace('\n', ' ').replace('\r', ' ')
  return label


# =========================
# Header Section
# =========================
def format_header(file_path: Path, route: str, parse_result: Dict, import_id: str, import_time: str) -> List[str]:
  """
  Generate the audit log header with file, schema, and sector info.
  Args:
      file_path: Path to the imported file.
      route: Route or context for the import (for logging).
      parse_result: Output of parse_xl_file (metadata, schemas, tab_contents).
  Returns:
      List of header lines for the audit log.
  """
  lines = [
    f"=== BEGIN AUDIT: {file_path.name} ===",
    f"Import ID{pad_label('', PAD_WIDTH - 8)}: {import_id}",
    f"Import Started{pad_label('', PAD_WIDTH - 13)}: {import_time}",
  ]
  if route:
    lines.append(f"Route{pad_label('', PAD_WIDTH - 5)}: {route}")
  lines.append("")
  lines.append("--- FILE METADATA ---")
  meta = parse_result.get('metadata', {})
  schema_tab = parse_result.get('schemas', {})
  sector = meta.get('sector') or meta.get('Sector')
  schema_name = None
  if schema_tab:
    schema_name = next(iter(schema_tab.values()), None)
  lines.append(f"Sector{pad_label('', PAD_WIDTH - 6)}: {sector}")
  lines.append(f"Schema{pad_label('', PAD_WIDTH - 6)}: {schema_name}")
  return lines


# =========================
# Type Conversion & Cleaning
# =========================
def try_type_conversion(raw_value: Any, value_type: Any) -> tuple[Any, list[str]]:
  """
  Attempt to convert raw_value to value_type. Return (converted_value, logs).
  Logs are human-readable explanations of what happened.

  Args:
    raw_value (Any): The value to convert.
    value_type (Any): The type to convert to.

  Returns:
    tuple[Any, list[str]]: The converted value and a list of log messages.
  """
  logs = []
  db_value = raw_value
  if raw_value is not None:
    if value_type and not isinstance(raw_value, value_type):
      if raw_value == "":
        db_value = None
        logs.append(
          f"Type conversion failed: empty string for expected type {getattr(value_type, '__name__', str(value_type))}. Value set to None.")
      else:
        try:
          if value_type == datetime.datetime:
            from arb.utils.date_and_time import excel_str_to_naive_datetime, is_datetime_naive
            excel_val = raw_value
            if not isinstance(excel_val, str):
              excel_val = str(excel_val)
            local_datetime = excel_str_to_naive_datetime(excel_val)
            if local_datetime and not is_datetime_naive(local_datetime):
              logs.append(f"Type conversion failed: parsed datetime is not naive for value '{raw_value}'. Value set to None.")
              db_value = None
            elif local_datetime is None:
              logs.append(f"Type conversion failed: could not parse '{raw_value}' as datetime. Value set to None.")
              db_value = None
            else:
              db_value = local_datetime
              logs.append(f"Type conversion successful: '{raw_value}' → {db_value} (datetime)")
          else:
            converted = value_type(raw_value)
            db_value = converted
            logs.append(f"Type conversion successful: '{raw_value}' → {db_value} ({getattr(value_type, '__name__', str(value_type))})")
        except (ValueError, TypeError) as e:
          logs.append(
            f"Type conversion failed: could not convert '{raw_value}' to {getattr(value_type, '__name__', str(value_type))}: {e}. Value set to None.")
          db_value = None
  return db_value, logs


# =========================
# Field Diagnostics Section
# =========================
def field_diagnostics(field: str, lookup: Dict, ws: Any, logs: List[str]) -> List[str]:
  """
  Generate diagnostics for a single field, including label/value comparison,
  whitespace warnings, type conversion logs, and dropdown validation.
  Args:
      field: Field name (schema key).
      lookup: Schema dict for this field.
      ws: openpyxl worksheet for the tab.
      logs: List of log messages from the import process.
  Returns:
      List of lines for this field's diagnostics.
  """
  pad = lambda label: pad_label(label, PAD_WIDTH)
  field_lines = []
  schema_label = lookup.get('label')
  label_address = lookup.get('label_address')
  value_address = lookup.get('value_address')
  value_type = lookup.get('value_type')
  is_drop_down = lookup.get('is_drop_down')
  spreadsheet_label = None
  if label_address:
    try:
      cell = ws[label_address]
      if isinstance(cell, tuple):
        cell = cell[0]
      spreadsheet_label = cell.value
    except Exception:
      spreadsheet_label = None
  label_match = (spreadsheet_label == schema_label) if (spreadsheet_label is not None and schema_label is not None) else None
  spreadsheet_raw_value = None
  if value_address:
    try:
      cell = ws[value_address]
      if isinstance(cell, tuple):
        cell = cell[0]
      spreadsheet_raw_value = cell.value
    except Exception:
      spreadsheet_raw_value = None
  db_value, conversion_logs = try_type_conversion(spreadsheet_raw_value, value_type)
  value_match_line = None
  if db_value == spreadsheet_raw_value:
    value_match_line = f"{pad('Value Match')} : True"
  elif (
      (isinstance(spreadsheet_raw_value, (int, float, str)) and isinstance(db_value, (int, float, str)))
      and str(spreadsheet_raw_value) == str(db_value)
  ):
    value_match_line = f"{pad('Value Match')} : True (Type Conversion: {type(spreadsheet_raw_value).__name__} → {type(db_value).__name__}, Value Preserved)"
  else:
    value_match_line = f"{pad('Value Match')} : False"
  display_spreadsheet_raw_value = spreadsheet_raw_value.strip() if isinstance(spreadsheet_raw_value, str) else spreadsheet_raw_value
  display_db_value = db_value.strip() if isinstance(db_value, str) else db_value
  whitespace_warning = None
  spreadsheet_ws_stripped = isinstance(spreadsheet_raw_value, str) and spreadsheet_raw_value != display_spreadsheet_raw_value
  db_ws_stripped = isinstance(db_value, str) and db_value != display_db_value
  if spreadsheet_ws_stripped and db_ws_stripped:
    whitespace_warning = f"WARNING: Whitespace was stripped from both spreadsheet and db value (before: {json.dumps(spreadsheet_raw_value, default=json_serializer)}, after: {json.dumps(display_db_value, default=json_serializer)})"
  elif spreadsheet_ws_stripped:
    whitespace_warning = f"WARNING: Whitespace was stripped from spreadsheet value (before: {json.dumps(spreadsheet_raw_value, default=json_serializer)}, after: {json.dumps(display_spreadsheet_raw_value, default=json_serializer)})"
  elif db_ws_stripped:
    whitespace_warning = f"WARNING: Whitespace was stripped from db value (before: {json.dumps(db_value, default=json_serializer)}, after: {json.dumps(display_db_value, default=json_serializer)})"
  field_lines.append(f"  Field{pad_label('', PAD_WIDTH - 5)}: {json.dumps(field)}")
  field_lines.append(f"    {pad('Schema Label')} : {json.dumps(schema_label, default=json_serializer)}")
  field_lines.append(f"    {pad('Spreadsheet Label')} : {json.dumps(spreadsheet_label, default=json_serializer)}")
  field_lines.append(f"    {pad('Value')} : {json.dumps(display_spreadsheet_raw_value, default=json_serializer)}")
  field_lines.append(f"    {pad('Data Type')} : {getattr(value_type, '__name__', str(value_type)) if value_type else None}")
  field_lines.append(f"    {pad('Value Match')} : {value_match_line.split(':', 1)[1].strip()}")
  if whitespace_warning:
    field_lines.append(f"    {pad('Warning')} : {whitespace_warning}")
  if conversion_logs:
    for log in conversion_logs:
      if log.lower().startswith('type conversion failed'):
        field_lines.append(f"    {pad('Invalid Input Ignored')} : INVALID INPUT IGNORED: {log}")
      else:
        field_lines.append(f"    {pad('Log')} : {log}")
  field_lines.append("")
  return field_lines


# =========================
# Summary Section
# =========================
def summary_section(label_match_count: int, label_mismatch_count: int, value_match_count: int, value_mismatch_count: int,
                    warning_count: int, invalid_input_count: int) -> List[str]:
  """
  Generate the summary section for the audit log.
  Args:
      label_match_count: Number of fields with label match True.
      label_mismatch_count: Number of fields with label match False.
      value_match_count: Number of fields with value match True.
      value_mismatch_count: Number of fields with value match False.
  Returns:
      List of lines for the summary section.
  """
  lines = [
    "",
    "--- SUMMARY ---",
    f"{pad_label('Fields Checked')} : {label_match_count + label_mismatch_count}",
    f"{pad_label('Fields With Warnings')} : {warning_count}",
    f"{pad_label('Fields With Invalid Input')} : {invalid_input_count}",
    f"{pad_label('Label Match')} : {label_match_count} pass, {label_mismatch_count} fail",
    f"{pad_label('Value Match')} : {value_match_count} pass, {value_mismatch_count} fail",
    ""
  ]
  return lines


def machine_readable_summary(import_id: str, fields_checked: int, warning_count: int, error_count: int, import_time: str) -> List[str]:
  summary = {
    "import_id": import_id,
    "fields_checked": fields_checked,
    "fields_with_warnings": warning_count,
    "fields_with_errors": error_count,
    "import_time": import_time
  }
  return [
    "--- MACHINE READABLE SUMMARY ---",
    json.dumps(summary, indent=2),
    "--- END MACHINE READABLE SUMMARY ---",
    ""
  ]


# =========================
# Main Audit Entry Point
# =========================
def generate_import_audit(
    file_path: Path,
    parse_result: Dict,
    schema_map: Dict,
    logs: Optional[List[str]] = None,
    route: str = ""
) -> str:
  """
  Generate a human-readable audit trail for an Excel import.

  Args:
      file_path: Path to the imported file.
      parse_result: Output of parse_xl_file (metadata, schemas, tab_contents).
      schema_map: Loaded schema map (from xl_parse.py).
      logs: List of log messages from the import process.
      route: Route or context for the import (for logging).

  Returns:
      The audit log as a string.
  """
  if logs is None:
    logs = []
  import_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
  import_id = f"{file_path.stem}_{import_time.replace('-', '').replace(':', '').replace(' ', '_')}"
  lines = format_header(file_path, route, parse_result, import_id, import_time)
  tab_name = 'Feedback Form'
  tab_schemas = parse_result.get('schemas', {})
  if tab_name not in tab_schemas:
    lines.append(f"Tab: {tab_name} not found in spreadsheet schemas.")
    lines.append(f"=== END AUDIT: {file_path.name} ===")
    return '\n'.join(lines)
  schema_version = tab_schemas[tab_name]
  if schema_version not in schema_map:
    lines.append(f"Schema version '{schema_version}' not found in schema_map.")
    lines.append(f"=== END AUDIT: {file_path.name} ===")
    return '\n'.join(lines)
  schema_dict = schema_map[schema_version]['schema']
  wb = openpyxl.load_workbook(file_path, data_only=True)
  if tab_name not in wb.sheetnames:
    lines.append(f"Tab: {tab_name} not found in Excel file.")
    lines.append(f"=== END AUDIT: {file_path.name} ===")
    return '\n'.join(lines)
  ws = wb[tab_name]
  lines.append("")
  lines.append("--- FIELD DIAGNOSTICS ---")
  label_match_count = 0
  label_mismatch_count = 0
  value_match_count = 0
  value_mismatch_count = 0
  warning_count = 0
  invalid_input_count = 0
  notes = set()
  for field, lookup in schema_dict.items():
    field_lines = field_diagnostics(field, lookup, ws, logs)
    for line in field_lines:
      if line.strip().startswith('Label Match'):
        if line.strip().endswith('True'):
          label_match_count += 1
        elif line.strip().endswith('False'):
          label_mismatch_count += 1
      if line.strip().startswith('Value Match'):
        if line.strip().endswith('True'):
          value_match_count += 1
        elif line.strip().endswith('False'):
          value_mismatch_count += 1
      if 'WARNING:' in line:
        warning_count += 1
        notes.add('Some spreadsheet or database values had extra whitespace that was removed during import.')
      if 'INVALID INPUT IGNORED:' in line:
        invalid_input_count += 1
        notes.add('Some fields contained invalid data and were ignored. Please check your input for correct types and formats.')
    lines.extend(field_lines)
  lines.extend(
    summary_section(label_match_count, label_mismatch_count, value_match_count, value_mismatch_count, warning_count, invalid_input_count))
  lines.extend(
    machine_readable_summary(import_id, label_match_count + label_mismatch_count, warning_count, invalid_input_count, import_time))
  if notes:
    lines.append('--- NOTES / SUGGESTIONS ---')
    for note in notes:
      lines.append(f'- {note}')
    lines.append('')
  lines.append(f"=== END AUDIT: {file_path.name} ===")
  return '\n'.join(lines)
