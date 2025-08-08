"""
excel_compare.py

Limitations (openpyxl and Python-based Excel comparison):
--------------------------------------------------------
- openpyxl does not fully capture all Excel formatting differences, especially:
    * Theme, tint, and indexed colors may not be resolved to their true RGB values.
    * Some border styles, conditional formatting, and advanced cell styles may not be detected.
    * Data validation, dropdowns, and protection flags are only partially supported.
    * Visual differences in Excel (as seen in the UI) may not be reflected in the Python object model.
    * Some workbook/sheet protection settings and cell-level protection may not be fully compared.
- As a result, two files that look different in Excel may appear identical to this script, and vice versa.
- For full-fidelity Excel comparison (including all formatting, protection, and UI-visible differences),
  consider using the provided VBA-based tool (see excel_compare_vba.bas).

This script is best used for basic content and common formatting checks, not for pixel-perfect or compliance-critical audits.

"""

import hashlib
from datetime import datetime
from pathlib import Path
from typing import Union

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

# Font and formatting attribute categories used for comparison
COMMON_FONT_ATTRS = ["bold", "italic", "size", "name"]
FULL_FONT_ATTRS = [
  "bold", "italic", "underline", "strike", "size", "name", "color", "vertAlign",
  "charset", "scheme", "family", "outline", "shadow"
]
FILL_ATTRS = ["patternType", "fgColor", "bgColor"]
ALIGNMENT_ATTRS = [
  "horizontal", "vertical", "wrapText", "shrinkToFit", "indent",
  "readingOrder", "textRotation"
]
PROTECTION_ATTRS = ["locked", "hidden"]
BORDER_SIDES = ["left", "right", "top", "bottom", "diagonal"]


def stringify_formula(cell) -> str:
  """Return a human-readable formula string or raw value for a cell."""
  try:
    if isinstance(cell.value, ArrayFormula):
      return cell.value.formula
    if cell.data_type == 'f':
      return str(cell.formula or cell._value or cell.value)
    return str(cell.value)
  except Exception as e:
    return f"[unreadable: {e}]"


def compute_sha256(path: Path) -> str:
  """Compute the SHA-256 hash of a file to detect binary-level equality."""
  with open(path, "rb") as f:
    return hashlib.sha256(f.read()).hexdigest()


def extract_data_validation_map(ws) -> dict:
  """
  Extract data validation rules from a worksheet and return a mapping
  from cell coordinate to validation settings.

  Returns:
    dict[str, dict]: Maps cell coordinate to a dict of validation attributes.
  """
  result = {}
  for dv in ws.data_validations.dataValidation:
    for cell_range in dv.cells:
      for cell in cell_range.cells:
        coord = ws.cell(row=cell.row, column=cell.col_idx).coordinate
        result[coord] = {
          "type": dv.type,
          "formula1": dv.formula1,
          "formula2": dv.formula2,
          "allow_blank": dv.allow_blank,
          "showDropDown": dv.showDropDown,
          "operator": dv.operator,
        }
  return result


def compare_workbook_protection(wb1, wb2) -> list[str]:
  """
  Compare top-level workbook protection settings.

  Returns:
    list[str]: Differences under [Workbook Protection Differences].
  """
  sec1 = getattr(wb1.security, "lockStructure", None), getattr(wb1.security, "lockWindows", None)
  sec2 = getattr(wb2.security, "lockStructure", None), getattr(wb2.security, "lockWindows", None)

  diffs = []
  if sec1[0] != sec2[0]:
    diffs.append(f"  lockStructure:\n    A: {sec1[0]}\n    B: {sec2[0]}")
  if sec1[1] != sec2[1]:
    diffs.append(f"  lockWindows:\n    A: {sec1[1]}\n    B: {sec2[1]}")
  return diffs


def compare_sheet_protection(ws1, ws2, sheet_name: str) -> list[str]:
  """
  Compare protection flags between two sheets.

  Returns:
    list[str]: Differences under [Sheet Protection Differences] per sheet.
  """
  attrs = [
    "sheet", "objects", "scenarios", "formatCells", "formatColumns", "formatRows",
    "insertColumns", "insertRows", "insertHyperlinks", "deleteColumns",
    "deleteRows", "selectLockedCells", "selectUnlockedCells", "sort", "autoFilter", "pivotTables"
  ]
  diffs = []
  for attr in attrs:
    v1 = getattr(ws1.protection, attr, None)
    v2 = getattr(ws2.protection, attr, None)
    if v1 != v2:
      diffs.append(f"  {attr}:\n    A: {v1}\n    B: {v2}")
  if diffs:
    return [f"[Sheet Protection Differences] ({sheet_name})"] + diffs
  return []


def compare_excel_content(path1: Path, path2: Path, formatting_mode: str = "common") -> list[str]:
  """
  Compare cell content, formulas, formatting, comments, dropdowns, and protection.

  Args:
    path1 (Path): Path to first Excel file.
    path2 (Path): Path to second Excel file.
    formatting_mode (str): 'off', 'common', or 'full'.

  Returns:
    list[str]: All differences grouped by category and sheet.
  """
  wb1 = load_workbook(path1, data_only=False)
  wb2 = load_workbook(path2, data_only=False)
  output = []

  # Workbook-level protection comparison
  workbook_protection_diffs = compare_workbook_protection(wb1, wb2)
  if workbook_protection_diffs:
    output.append("[Workbook Protection Differences]")
    output.extend(workbook_protection_diffs)

  sheets1 = set(wb1.sheetnames)
  sheets2 = set(wb2.sheetnames)
  only_in_a = sorted(sheets1 - sheets2)
  only_in_b = sorted(sheets2 - sheets1)
  in_both = sorted(sheets1 & sheets2)

  if only_in_a:
    output.append("[Sheets only in A]")
    for name in only_in_a:
      output.append(f"  {name}")
  if only_in_b:
    output.append("[Sheets only in B]")
    for name in only_in_b:
      output.append(f"  {name}")

  for sheet in in_both:
    output.append(f"\n=== Sheet: {sheet} ===")
    ws1 = wb1[sheet]
    ws2 = wb2[sheet]
    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    sheet_protection_diffs = compare_sheet_protection(ws1, ws2, sheet)
    if sheet_protection_diffs:
      output.extend(sheet_protection_diffs)

    content_diffs, formula_diffs = [], []
    comment_diffs, formatting_diffs, validation_diffs, protection_diffs = [], [], [], []
    validations1 = extract_data_validation_map(ws1)
    validations2 = extract_data_validation_map(ws2)

    for row in range(1, max_row + 1):
      for col in range(1, max_col + 1):
        c1 = ws1.cell(row=row, column=col)
        c2 = ws2.cell(row=row, column=col)
        coord = c1.coordinate

        v1 = stringify_formula(c1)
        v2 = stringify_formula(c2)
        if v1 != v2:
          content_diffs.append(f"  {coord}:\n    A: {v1}\n    B: {v2}")

        if c1.data_type == 'f' or c2.data_type == 'f':
          if v1 != v2:
            formula_diffs.append(f"  {coord}:\n    A: {v1}\n    B: {v2}")

        comm1 = c1.comment.text if c1.comment else None
        comm2 = c2.comment.text if c2.comment else None
        if comm1 != comm2:
          comment_diffs.append(f"  {coord}:\n    A: {comm1}\n    B: {comm2}")

        # Protection attributes always checked
        for attr in PROTECTION_ATTRS:
          if getattr(c1.protection, attr) != getattr(c2.protection, attr):
            protection_diffs.append(f"  {coord}: protection.{attr} changed")

        if formatting_mode != "off":
          for attr in (COMMON_FONT_ATTRS if formatting_mode == "common" else FULL_FONT_ATTRS):
            if getattr(c1.font, attr) != getattr(c2.font, attr):
              formatting_diffs.append(f"  {coord}: font.{attr} changed")
          for attr in FILL_ATTRS:
            if getattr(c1.fill, attr) != getattr(c2.fill, attr):
              formatting_diffs.append(f"  {coord}: fill.{attr} changed")
          for attr in ALIGNMENT_ATTRS:
            if getattr(c1.alignment, attr) != getattr(c2.alignment, attr):
              formatting_diffs.append(f"  {coord}: alignment.{attr} changed")
          if c1.number_format != c2.number_format:
            formatting_diffs.append(
              f"  {coord}: number_format changed\n    A: {c1.number_format}\n    B: {c2.number_format}")
          if formatting_mode == "full":
            for side in BORDER_SIDES:
              b1 = getattr(c1.border, side)
              b2 = getattr(c2.border, side)
              if b1.style != b2.style or b1.color != b2.color:
                formatting_diffs.append(f"  {coord}: border.{side} changed")

        val1 = validations1.get(coord)
        val2 = validations2.get(coord)
        if val1 != val2:
          validation_diffs.append(f"  {coord}:\n    A: {val1}\n    B: {val2}")

    if content_diffs:
      output.append("[Content Differences]")
      output.extend(content_diffs)
    if formula_diffs:
      output.append("[Formula Differences]")
      output.extend(formula_diffs)
    if comment_diffs:
      output.append("[Comment Differences]")
      output.extend(comment_diffs)
    if validation_diffs:
      output.append("[Dropdown/Data Validation Differences]")
      output.extend(validation_diffs)
    if protection_diffs:
      output.append("[Cell Protection Differences]")
      output.extend(protection_diffs)
    if formatting_diffs:
      output.append("[Formatting Differences]")
      output.extend(formatting_diffs)

  return output


def compare_excel_files(file_a_path: Union[str, Path],
                        file_b_path: Union[str, Path],
                        formatting_mode: str = "common",
                        log_to_file: bool = False) -> list[str]:
  """
  Compare two Excel workbooks at the binary and content level.

  Args:
    file_a_path (str | Path): Path to file A.
    file_b_path (str | Path): Path to file B.
    formatting_mode (str): Formatting check mode: 'off', 'common', or 'full'.
    log_to_file (bool): If true, then log to file

  Returns:
    list[str]: List of human-readable comparison results.
  """
  file_a = Path(file_a_path)
  file_b = Path(file_b_path)
  output = [f"Comparing:\n  A: {file_a.name}\n  B: {file_b.name}\n"]

  hash_a = compute_sha256(file_a)
  hash_b = compute_sha256(file_b)
  output.append("SHA-256 Hashes:")
  output.append(f"  A: {hash_a}")
  output.append(f"  B: {hash_b}")

  if hash_a == hash_b:
    output.append("\u2714 Files are identical at the binary level.\n")
    return output

  output.append("\u2192 Hashes differ; comparing content...\n")
  output.extend(compare_excel_content(file_a, file_b, formatting_mode=formatting_mode))

  if log_to_file:
    now = datetime.now()
    out_path = Path(f"comparison_at_{now.strftime('%Y%m%d_%H%M%S')}.txt")
    print(f"Creating comparison file: {out_path}")

    with open(out_path, "w", encoding="utf-8") as f:
      f.write(f"Comparing Excel files \n\t A: {file_a} \n\t B: {file_b}\n\n")
      for line in output:
        f.write(line + "\n")
  return output


def compare_excel_directories(dir_a: Path, dir_b: Path, formatting_mode: str = "common") -> None:
  """
  Compare all Excel files in two directories by name, producing a timestamped report.

  Args:
    dir_a (Path): Path to first directory.
    dir_b (Path): Path to second directory.
    formatting_mode (str): Formatting check mode: 'off', 'common', or 'full'.
  """
  now = datetime.now()
  out_path = Path(f"comparison_at_{now.strftime('%Y%m%d_%H%M%S')}.txt")
  print(f"Creating comparison file: {out_path}")

  excel_files_a = {f.name: f for f in dir_a.glob("*.xlsx") if f.is_file()}
  excel_files_b = {f.name: f for f in dir_b.glob("*.xlsx") if f.is_file()}

  only_in_a = sorted(set(excel_files_a) - set(excel_files_b))
  only_in_b = sorted(set(excel_files_b) - set(excel_files_a))
  in_both = sorted(set(excel_files_a) & set(excel_files_b))

  with open(out_path, "w", encoding="utf-8") as f:
    f.write(f"Comparing Excel files in:\n  A: {dir_a}\n  B: {dir_b}\n\n")
    f.write(f"Files only in A ({len(only_in_a)}):\n")
    for name in only_in_a:
      f.write(f"  {name}\n")
    f.write(f"\nFiles only in B ({len(only_in_b)}):\n")
    for name in only_in_b:
      f.write(f"  {name}\n")
    f.write(f"\nFiles in both ({len(in_both)}):\n")

    for name in in_both:
      f.write(f"\n=== Comparing {name} ===\n")
      result = compare_excel_files(
        excel_files_a[name],
        excel_files_b[name],
        formatting_mode=formatting_mode
      )
      for line in result:
        f.write(line + "\n")

  print(f"Comparison complete. Output saved to: {out_path.resolve()}")


if __name__ == "__main__":
  # todo - i don't think this detects changes in the date string formatting, which should be implemented

  # local_path = Path(r"C:\tony_local\pycharm\feedback_portal\feedback_forms\processed_versions\xl_workbooks")
  local_path = Path(r"C:\tony_local\pycharm\feedback_portal\feedback_forms\current_versions")
  sharepoint_path = Path(
    r"C:\Users\theld\OneDrive - California Air Resources Board\OneDriveLinks\RD Satellite Project - Operator Notification Materials for Review\spreadsheets\xl_workbooks")

  if True:
    compare_excel_directories(local_path, sharepoint_path, formatting_mode="full")

  if True:
    # local_path = Path(r"C:\tony_local\pycharm\feedback_portal\diagnostics")
    local_path = Path(r"D:\local\cursor\feedback_portal\diagnostics")
    file_a = local_path / "dairy_digester_operator_feedback_v006_for_review_local.xlsx"
    file_b = local_path / "dairy_digester_operator_feedback_v006_for_review_sharepoint.xlsx"
    print(f"{file_a}\n{file_b}")
    compare_excel_files(file_a, file_b, formatting_mode="full", log_to_file=True)
