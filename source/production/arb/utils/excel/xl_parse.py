"""
Module to parse and ingest Excel spreadsheet contents.

This module provides logic to convert Excel forms into structured dictionary representations,
including extraction of tab contents, metadata, and schema references. It is primarily used
to support automated feedback template parsing.

Notes:
  - `schema_file_map` is a dict where keys are schema names and values are paths to JSON files.
  - `schema_map` is a dict where keys are schema names and values are:
      {"schema": schema_dict, "metadata": metadata_dict}

Example:
  Input : xl_schema_map['oil_and_gas_v03']['schema']
  Output: Dictionary representing the oil and gas schema
"""

import copy
import datetime
import logging
from pathlib import Path

import openpyxl

from arb.portal.constants import PLEASE_SELECT
from arb.utils.date_and_time import excel_str_to_naive_datetime, is_datetime_naive
from arb.utils.excel.xl_file_structure import PROCESSED_VERSIONS
from arb.utils.json import json_load_with_meta, json_save_with_meta
from arb.utils.misc import sanitize_for_utf8

logger = logging.getLogger(__name__)

# Spreadsheet formatting constants
EXCEL_SCHEMA_TAB_NAME = '_json_schema'
EXCEL_METADATA_TAB_NAME = '_json_metadata'
EXCEL_TOP_LEFT_KEY_VALUE_CELL = '$B$15'

# --- Schema aliasing for backward compatibility ---
# Maps old schema names to new schema names {old_schema_name: new_schema_name}.
# If a file references an outdated schema,
# the parser will use the mapped schema and log a warning.
schema_alias = {
  "energy_v00_01": "energy_v01_00",
  "generic_v00_01": "generic_v01_00",
}

# xl_schema_map based on Excel PROCESSED_VERSIONS files
xl_schema_file_map = {}  # type: dict[str, Path]
xl_schema_map = {}  # type: dict[str, dict]


def initialize_module() -> None:
  """
  Initialize the module by calling set_globals().

  This function loads default schema mappings into global variables.
  """
  logger.debug(f"initialize_module() called")
  set_globals()


def set_globals(xl_schema_file_map_: dict[str, Path] | None = None) -> None:
  """
  Set module-level global variables for schema file map and loaded schema map.

  Args:
    xl_schema_file_map_ (dict[str, Path] | None): Optional override for the schema file map.
      If not provided, use a default list of pre-defined schema files based on TEMPLATES.

  Notes:
    - Calls `load_schema_file_map()` to populate xl_schema_map from JSON files.
    - Uses TEMPLATES structure from xl_create.py for consistency.
  """
  global xl_schema_file_map, xl_schema_map

  logger.debug(f"set_globals() called with {xl_schema_file_map_=}")

  if xl_schema_file_map_ is None:
    # Import TEMPLATES from xl_create to ensure consistency
    from arb.utils.excel.xl_hardcoded import EXCEL_TEMPLATES

    xl_schema_file_map = {}
    for template in EXCEL_TEMPLATES:
      schema_version = template["schema_version"]
      schema_path = PROCESSED_VERSIONS / "xl_schemas" / f"{schema_version}.json"
      xl_schema_file_map[schema_version] = schema_path
  else:
    xl_schema_file_map = xl_schema_file_map_

  # load all the schemas
  if xl_schema_file_map:
    xl_schema_map = load_schema_file_map(xl_schema_file_map)

  logger.debug(f"globals are now: {xl_schema_file_map=}, {xl_schema_map=}")


def load_schema_file_map(schema_file_map: dict[str, Path]) -> dict[str, dict]:
  """
  Load JSON schema and metadata from a mapping of schema name to a file path.

  Args:
    schema_file_map (dict[str, Path]): Keys are schema names, values are JSON schema file paths.

  Returns:
    dict[str, dict]: Dictionary where keys are schema names and values are dicts with:
      - "schema": The schema dictionary.
      - "metadata": Metadata extracted from the JSON.
  """
  logger.debug(f"load_schema_file_map() called with {schema_file_map=}")

  schema_map = {}

  for schema_name, json_path in schema_file_map.items():
    schema, metadata = json_load_with_meta(json_path)
    schema_map[schema_name] = {"schema": schema, "metadata": metadata}

  return schema_map


def create_schema_file_map(schema_path: str | Path | None = None,
                           schema_names: list[str] | None = None) -> dict[str, Path]:
  """
  Create a dictionary mapping schema names to their JSON file paths.

  Args:
    schema_path (str | Path | None): Folder containing schema files. Defaults to processed versions dir.
    schema_names (list[str] | None): Names of schemas to include. Defaults to schemas from TEMPLATES.

  Returns:
    dict[str, Path]: Map from schema name to a schema file path.
  """
  logger.debug(f"create_schema_file_map() called with {schema_path=}, {schema_names=}")

  if isinstance(schema_path, str):
    schema_path = Path(schema_path)

  if schema_path is None:
    schema_path = PROCESSED_VERSIONS / "xl_schemas"
  if schema_names is None:
    # Import TEMPLATES from xl_create to ensure consistency
    from arb.utils.excel.xl_hardcoded import EXCEL_TEMPLATES
    schema_names = [template["schema_version"] for template in EXCEL_TEMPLATES]

  schema_file_map = {}

  for schema_name in schema_names:
    schema_file_name = schema_name + ".json"
    schema_file_path = schema_path / schema_file_name
    schema_file_map[schema_name] = schema_file_path

  return schema_file_map


def load_xl_schema(file_name: str | Path) -> tuple[dict, dict]:
  """
  Load schema and metadata from a JSON file.

  Args:
    file_name (str | Path): Path to a JSON schema file.

  Returns:
    tuple[dict, dict]: Tuple of (schema dict, metadata dict).
  """
  logger.debug(f"load_xl_schema() called with {file_name=}")
  schema, metadata = json_load_with_meta(file_name)
  return schema, metadata


def parse_xl_file(xl_path: str | Path,
                  schema_map: dict[str, dict] | None = None) -> dict[str, dict]:
  """
  Parse a spreadsheet and return a dictionary representation using the given schema.

  Args:
    xl_path (str | Path): Path to the Excel spreadsheet.
    schema_map (dict[str, dict] | None): Map of schema names to their definitions.

  Returns:
    dict: Dictionary with extracted metadata, schemas, and tab contents.

  Notes:
  - tutorial on openpyxl: https://openpyxl.readthedocs.io/en/stable/tutorial.html
  """
  logger.debug(f"parse_xl_with_schema_dict() called with {xl_path=}, {schema_map=}")

  if schema_map is None:
    schema_map = xl_schema_map

  # Dictionary data structure to store Excel contents
  result = {}
  result['metadata'] = {}
  result['schemas'] = {}
  result['tab_contents'] = {}

  # Notes on data_only argument.  By default, .value returns the 'formula' in the cell.
  # If data_only=True, then .value returns the last 'value' that was evaluated at the cell.
  wb = openpyxl.load_workbook(xl_path, keep_vba=False, data_only=True)

  # Extract metadata and schema information from hidden tabs
  if EXCEL_METADATA_TAB_NAME in wb.sheetnames:
    logger.debug(f"metadata tab detected in Excel file")
    result['metadata'] = get_spreadsheet_key_value_pairs(wb,
                                                         EXCEL_METADATA_TAB_NAME,
                                                         EXCEL_TOP_LEFT_KEY_VALUE_CELL
                                                         )
    # todo - maybe want to alias the Sector here?  Refinery -> Energy
  if EXCEL_SCHEMA_TAB_NAME in wb.sheetnames:
    logger.debug(f"Schema tab detected in Excel file")
    result['schemas'] = get_spreadsheet_key_value_pairs(wb,
                                                        EXCEL_SCHEMA_TAB_NAME,
                                                        EXCEL_TOP_LEFT_KEY_VALUE_CELL)
    # todo - maybe want to alias schemas here before extract tabs
  else:
    ValueError(f'Spreadsheet must have a {EXCEL_SCHEMA_TAB_NAME} tab')

  # extract data tabs content using specified schemas
  new_result = extract_tabs(wb, schema_map, result)

  return new_result


def extract_tabs(wb: openpyxl.Workbook,
                 schema_map: dict[str, dict],
                 xl_as_dict: dict) -> dict:
  """
  Extract data from the data tabs that are enumerated in the schema tab.

  Args:
    wb (Workbook): OpenPyXL workbook object.
    schema_map (dict[str, dict]): Schema map with schema definitions.
    xl_as_dict (dict): Parsed Excel content, including 'schemas' and 'metadata'.
                       Dictionary with schema tab where keys are the data tab names and values are the formatting_schema to
                       parse the tab.

  Returns:
    dict: Updated xl_as_dict including parsed 'tab_contents'.
  """

  skip_please_selects = False

  result = copy.deepcopy(xl_as_dict)

  for tab_name, formatting_schema in result['schemas'].items():
    resolved_schema = ensure_schema(formatting_schema, schema_map, schema_alias, logger)
    if not resolved_schema:
      continue
    logger.debug(f"Extracting data from '{tab_name}', using the formatting schema '{formatting_schema}'")
    result['tab_contents'][tab_name] = {}

    ws = wb[tab_name]
    format_dict = schema_map[resolved_schema]['schema']

    for html_field_name, lookup in format_dict.items():
      value_address = lookup['value_address']
      value_type = lookup['value_type']
      is_drop_down = lookup['is_drop_down']
      # works, but you get a lint error because this will break if value_address is not a single cell
      value = ws[value_address].value  # type: ignore[attr-defined]
      # Strip whitespace for string fields and log if changed
      if value is not None and value_type == str and isinstance(value, str):
        # todo - use sanitize_for_logging here?
        value = sanitize_for_utf8(value)
        stripped_value = value.strip()
        if value != stripped_value:
          logger.warning(
            f"Whitespace detected for field '{html_field_name}' at {value_address}: before strip: {repr(value)}, after strip: {repr(stripped_value)}")
        value = stripped_value

      if skip_please_selects is True:
        if is_drop_down and value == PLEASE_SELECT:
          logger.debug(f"Skipping {html_field_name} because it is a drop down and is set to {PLEASE_SELECT}")
          continue

      # Try to cast the spreadsheet data to the desired type if possible
      if value is not None:
        if not isinstance(value, value_type):
          # if it is not supposed to be of type string, but it is a zero-length string, turn it to None
          if value == "":
            value = None
          else:
            logger.warning(f"Warning: <{html_field_name}> value at <{lookup['value_address']}> is <{value}> "
                           f"and is of type <{type(value)}> whereas it should be of type <{value_type}>.  "
                           f"Attempting to convert the value to the correct type")
            try:
              # convert to datetime using a parser if possible
              # todo - datetime - seems like I could cast to utc here for persistence
              if value_type == datetime.datetime:
                local_datetime = excel_str_to_naive_datetime(value)
                if local_datetime and not is_datetime_naive(local_datetime):
                  logger.warning(f"Date time {value} is not a naive datetime, skipping to avoid data corruption")
                  continue
                value = local_datetime
              else:
                # Use default repr-like conversion if not a datetime
                value = value_type(value)
              logger.info(f"Type conversion successful.  value is now <{value}> with type: <{type(value)}>")
            except (ValueError, TypeError) as e:
              logger.warning(f"Type conversion failed, resetting value to None")
              value = None

      result['tab_contents'][tab_name][html_field_name] = value

      if 'label_address' in lookup and 'label' in lookup:
        label_address = lookup['label_address']
        # works, but you get a lint error because this will break if value_address is not a single cell
        label_xl = ws[label_address].value  # type: ignore[attr-defined]
        label_schema = lookup['label']
        if label_xl != label_schema:
          logger.warning(f"Schema data label and spreadsheet data label differ."
                         f"\n\tschema label = {label_schema}\n\tspreadsheet label ({label_address}) = {label_xl}")

    logger.debug(f"Initial spreadsheet extraction of '{tab_name}' yields {result['tab_contents'][tab_name]}")
    # Some cells should be spit into multiple dictionary entries (such as full name, lat/log)
    split_compound_keys(result['tab_contents'][tab_name])

    logger.debug(f"Final corrected spreadsheet extraction of '{tab_name}' yields {result['tab_contents'][tab_name]}")

  return result


def ensure_schema(formatting_schema: str, schema_map: dict, schema_alias: dict, logger: logging.Logger) -> str | None:
  """
  Resolves a schema version using the schema map and alias mapping.
  Logs a warning if an alias is used. Returns the resolved schema version, or None if not found.

  Args:
      formatting_schema (str): The schema version to resolve.
      schema_map (dict): The mapping of valid schema versions.
      schema_alias (dict): The mapping of old schema names to new ones.
      logger: Logger for warnings/errors.

  Returns:
      str | None: The resolved schema version, or None if not found.
  """
  if formatting_schema in schema_map:
    return formatting_schema
  if formatting_schema in schema_alias:
    new_schema_version = schema_alias[formatting_schema]
    logger.warning(f"Schema '{formatting_schema}' not found. Using alias '{new_schema_version}' instead.")
    if new_schema_version in schema_map:
      return new_schema_version
    else:
      logger.error(f"Alias '{new_schema_version}' not found in schema_map.")
      return None
  logger.error(f"Schema '{formatting_schema}' not found and no alias available.")
  return None


def split_compound_keys(dict_: dict) -> None:
  """
  Decompose compound keys into atomic fields.

  Remove key/value pairs of entries that potentially contain compound keys and replace them with key value pairs
  that are more atomic.

  Args:
    dict_ (dict): Dictionary with potentially compound fields (e.g., lat_and_long).

  Raises:
    ValueError: If 'lat_and_long' is improperly formatted.
  """
  for html_field_name in list(dict_.keys()):
    value = dict_[html_field_name]

    if html_field_name == 'lat_and_long':
      if value:
        lat_longs = value.split(',')
        if len(lat_longs) == 2:
          dict_['lat_arb'] = lat_longs[0]
          dict_['long_arb'] = lat_longs[1]
        else:
          raise ValueError(f"Lat long must be a blank or a comma separated list of lat/long pairs")
      del dict_[html_field_name]


def get_spreadsheet_key_value_pairs(wb: openpyxl.Workbook,
                                    tab_name: str,
                                    top_left_cell: str) -> dict[str, str | None]:
  """
  Read key-value pairs from a worksheet starting at a given cell.

  Args:
    wb (Workbook): OpenPyXL workbook object.
    tab_name (str): Name of the worksheet tab.
    top_left_cell (str): Top-left cell of the key/value pair region.

  Returns:
    dict[str, str | None]: Parsed key-value pairs.
  """

  # logger.debug(f"{type(wb)=}, ")
  ws = wb[tab_name]

  # logger.debug(f"{type(ws)=}, ")

  return_dict = {}

  row_offset = 0
  while True:
    key = ws[top_left_cell].offset(row=row_offset).value
    value = ws[top_left_cell].offset(row=row_offset, column=1).value
    if key not in ["", None]:
      return_dict[key] = value
      row_offset += 1
    else:
      break

  return return_dict


def get_json_file_name_old(file_name: Path) -> Path | None:
  """
  Depreciated: use convert_upload_to_json instead

  Convert a file name (Excel or JSON) into a JSON file name, parsing if needed.

  Args:
    file_name (Path): The uploaded file.

  Returns:
    Path | None: JSON file path if parsed or detected, otherwise None.

  Notes:
  - If the file_name is a JSON file (has .json extension), the file_name is the json_file_name.
  - If the file is an Excel file (.xlsx extension), try to parse it into a JSON file and return the JSON file name
    of the parsed contents.
  - If the file is neither a json file nr a spreadsheet that can be parsed into a json file, return None.
  - If the file was already a json file, return the file name unaltered.
  - If the file was an Excel file, return the json file that its data was extracted to.
  - For all other file types return None.
  """

  json_file_name = None

  extension = file_name.suffix
  # logger.debug(f"{extension=}")
  if extension == ".xlsx":
    logger.debug(f"Excel file upload detected.  File name: {file_name}")
    # xl_as_dict = xl_path_to_dict(file_name)
    xl_as_dict = parse_xl_file(file_name)
    logger.debug(f"{xl_as_dict=}")
    json_file_name = file_name.with_suffix('.json')

    logger.debug(f"Saving extracted data from Excel as: {json_file_name}")
    json_save_with_meta(json_file_name, xl_as_dict)
  elif extension == ".json":
    logger.debug(f"Json file upload detected.  File name: {file_name}")
    json_file_name = Path(file_name)
  else:
    logger.debug(f"Unknown file type upload detected.  File name: {file_name}")

  return json_file_name


def convert_upload_to_json(file_path: Path) -> Path | None:
  """
  Convert an uploaded Excel or JSON file into a valid JSON payload file.

  Args:
    file_path (Path): Path to the uploaded file.

  Returns:
    Path | None:
      - Path to JSON file (either original or newly created), or
      - None if file type is unsupported or conversion fails.

  Behavior:
    - If the file has a `.json` extension:
        → Assume it is valid JSON and return as-is.
    - If the file has a `.xlsx` extension:
        → Attempt to parse using Excel schema logic.
        → Save converted contents as a `.json` file in the same directory.
        → Return the path to that JSON file.
    - If the file is neither `.xlsx` nor `.json`:
        → Log a warning and return None.

  Side Effects:
    - May write a `.json` file to disk if an Excel file is successfully parsed.
  """
  extension = file_path.suffix.lower()
  json_path = None

  if extension == ".xlsx":
    logger.debug(f"Excel upload detected: {file_path}")
    try:
      xl_as_dict = parse_xl_file(file_path)
      logger.debug(f"Parsed Excel to dict: {xl_as_dict.keys()}")
      json_path = file_path.with_suffix(".json")
      logger.debug(f"Saving Excel-derived JSON as: {json_path}")
      json_save_with_meta(json_path, xl_as_dict)
    except Exception as e:
      logger.warning(f"Excel parsing failed for {file_path}: {e}")
      return None

  elif extension == ".json":
    logger.debug(f"JSON upload detected: {file_path}")
    json_path = file_path

  else:
    logger.warning(f"Unsupported file type: {file_path}")

  return json_path


def test_load_schema_file_map() -> None:
  """
  Debug test for loading a schema file map and displaying contents.
  """
  logger.debug(f"test_load_schema_file_map() called")
  schema_map = create_schema_file_map()
  schemas = load_schema_file_map(schema_map)
  logger.debug(f"{schemas=}")


def test_load_xl_schemas() -> None:
  """
  Debug test for loading default schemas from xl_schema_file_map.
  """
  from arb.logging.arb_logging import get_pretty_printer
  _, pp_log = get_pretty_printer()

  logger.debug(f"Testing load_xl_schemas() with test_load_xl_schemas")
  schemas = load_schema_file_map(xl_schema_file_map)
  logger.debug(f"Testing load_xl_schemas() with test_load_xl_schemas")
  logging.debug(f"schemas = {pp_log(schemas)}")
  # logging.debug(f"\n schemas= \n{dict_to_str(schemas)}")


def test_parse_xl_file() -> None:
  """
  Debug test to parse a known Excel file into dictionary form using schemas.
  """
  logger.debug(f"test_parse_xl_file() called")
  xl_path = PROCESSED_VERSIONS / "xl_workbooks" / "landfill_operator_feedback_v070_populated_01.xlsx"
  print(f"{xl_path=}")
  result = parse_xl_file(xl_path, xl_schema_map)
  logger.debug(f"{result=}")


def main() -> None:
  """
  Run all schema and Excel file parsing test functions for diagnostic purposes.
  """
  test_load_xl_schemas()
  test_load_schema_file_map()
  test_parse_xl_file()
  # initialize_module()

# -----------------------------------------------------------------------------
# Initialize module global values
# -----------------------------------------------------------------------------
initialize_module()

if __name__ == "__main__":
  main()
