"""
Excel processing orchestration module.

This module provides the main processing orchestration for Excel files,
integrating validation, data extraction, and result generation into
a cohesive workflow.

Classes:
    ExcelProcessor: Main orchestrator for Excel processing operations

Author: AI Assistant
Created: 2025-01-27
Version: 1.0
"""

import logging
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.workbook.workbook import Workbook

from ..core import (
    ExcelParseConfig,
    ExcelParseResult,
    ProcessingStats,
    ValidationResult,
    ExcelProcessingError,
    ErrorCodes
)
from ..validation import ExcelValidator, SchemaValidator, DataValidator


class ExcelProcessor:
    """
    Main orchestrator for Excel processing operations.
    
    This class coordinates the entire Excel processing workflow:
    - File validation using ExcelValidator
    - Schema validation using SchemaValidator  
    - Data extraction and validation using DataValidator
    - Result generation and error handling
    
    The processor provides a high-level interface for processing
    Excel files with comprehensive validation and detailed reporting.
    
    Example:
        processor = ExcelProcessor()
        config = ExcelParseConfig(max_file_size_mb=50)
        
        # Process Excel file with validation
        result = processor.process_file(
            file_path=Path("sample.xlsx"),
            config=config,
            required_tabs=["Sheet1", "Sheet2"]
        )
        
        if result.success:
            print(f"Processing successful: {result.get_processing_summary()}")
        else:
            print(f"Processing failed: {result.get_error_summary()}")
    """
    
    def __init__(self, config: Optional[ExcelParseConfig] = None):
        """
        Initialize the Excel processor.
        
        Args:
            config: Configuration for processing behavior. If None, uses default settings.
        """
        self.config = config or ExcelParseConfig()
        self.logger = logging.getLogger(__name__)
        
        # Initialize validation components
        self.excel_validator = ExcelValidator(self.config)
        self.schema_validator = SchemaValidator()
        self.data_validator = DataValidator()
        
        # Processing statistics
        self.stats = ProcessingStats()
    
    def process_file(self, file_path: Path, 
                    required_tabs: Optional[List[str]] = None,
                    schemas: Optional[Dict[str, Dict[str, Any]]] = None) -> ExcelParseResult:
        """
        Process an Excel file with comprehensive validation and extraction.
        
        This method orchestrates the complete processing workflow:
        1. File validation (existence, format, size, structure)
        2. Schema validation (if schemas provided)
        3. Data extraction and validation
        4. Result generation with detailed reporting
        
        Args:
            file_path: Path to the Excel file to process
            required_tabs: Optional list of required worksheet names
            schemas: Optional dictionary of schemas by tab name
            
        Returns:
            ExcelParseResult: Comprehensive result of the processing operation
            
        Raises:
            ExcelProcessingError: If processing fails and strict mode is enabled
        """
        start_time = time.time()
        self.stats.start_time = start_time
        
        try:
            self.logger.info(f"Starting Excel file processing: {file_path}")
            
            # Step 1: File validation
            validation_results = self._validate_file(file_path, required_tabs)
            if not self._should_continue_processing(validation_results):
                return self._create_failed_result(
                    file_path, validation_results, start_time,
                    "File validation failed"
                )
            
            # Step 2: Load workbook
            workbook = self._load_workbook(file_path)
            if not workbook:
                return self._create_failed_result(
                    file_path, validation_results, start_time,
                    "Failed to load workbook"
                )
            
            # Step 3: Schema validation (if schemas provided)
            if schemas:
                schema_results = self._validate_schemas(schemas, workbook)
                validation_results.extend(schema_results)
                if not self._should_continue_processing(schema_results):
                    return self._create_failed_result(
                        file_path, validation_results, start_time,
                        "Schema validation failed"
                    )
            
            # Step 4: Data extraction
            extraction_results = self._extract_data(workbook, schemas)
            
            # Step 5: Generate final result
            result = self._create_success_result(
                file_path, validation_results, extraction_results, start_time
            )
            
            self.logger.info(f"Excel file processing completed successfully: {file_path}")
            return result
            
        except Exception as e:
            self.logger.error(f"Excel file processing failed: {file_path}, error: {str(e)}")
            return self._create_failed_result(
                file_path, validation_results if 'validation_results' in locals() else [],
                start_time, f"Processing error: {str(e)}"
            )
        finally:
            # Clean up workbook
            if 'workbook' in locals():
                try:
                    workbook.close()
                except Exception:
                    pass
    
    def process_tab(self, workbook: Workbook, tab_name: str,
                   schema: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Process a single tab/worksheet from a workbook.
        
        Args:
            workbook: OpenPyXL workbook object
            tab_name: Name of the tab to process
            schema: Optional schema for the tab
            
        Returns:
            Dictionary containing extracted data and validation results
        """
        try:
            if tab_name not in workbook.sheetnames:
                raise ExcelProcessingError(
                    f"Tab '{tab_name}' not found in workbook",
                    ErrorCodes.SCHEMA_NOT_FOUND,
                    context={"available_tabs": workbook.sheetnames}
                )
            
            worksheet = workbook[tab_name]
            
            # Extract data based on schema or use default extraction
            if schema:
                return self._extract_tab_with_schema(worksheet, schema)
            else:
                return self._extract_tab_default(worksheet)
                
        except Exception as e:
            self.logger.error(f"Tab processing failed: {tab_name}, error: {str(e)}")
            raise ExcelProcessingError(
                f"Failed to process tab '{tab_name}': {str(e)}",
                ErrorCodes.PROCESSING_FAILED,
                context={"tab_name": tab_name},
                original_exception=e
            )
    
    def _validate_file(self, file_path: Path, required_tabs: Optional[List[str]]) -> List[ValidationResult]:
        """
        Perform comprehensive file validation.
        
        Args:
            file_path: Path to the file to validate
            required_tabs: Optional list of required worksheet names
            
        Returns:
            List of validation results
        """
        self.logger.debug(f"Validating file: {file_path}")
        
        # Perform comprehensive validation
        validation_results = self.excel_validator.validate_file_comprehensive(
            file_path, required_tabs
        )
        
        # Log validation results
        for result in validation_results:
            if result.is_valid:
                self.logger.debug(f"Validation passed: {result.field_name}")
            else:
                self.logger.warning(f"Validation failed: {result.field_name}: {result.message}")
        
        return validation_results
    
    def _should_continue_processing(self, validation_results: List[ValidationResult]) -> bool:
        """
        Determine if processing should continue based on validation results.
        
        Args:
            validation_results: List of validation results
            
        Returns:
            True if processing should continue, False otherwise
        """
        if not validation_results:
            return True
        
        # Check if any critical validations failed
        critical_failures = [
            r for r in validation_results 
            if not r.is_valid and r.severity == "ERROR"
        ]
        
        if critical_failures and self.config.strict_mode:
            return False
        
        # Allow processing to continue if only warnings or non-critical errors
        return True
    
    def _load_workbook(self, file_path: Path) -> Optional[Workbook]:
        """
        Load the Excel workbook.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            OpenPyXL workbook object or None if loading fails
        """
        try:
            self.logger.debug(f"Loading workbook: {file_path}")
            
            # Load workbook with appropriate options
            workbook = openpyxl.load_workbook(
                file_path,
                read_only=True,  # Read-only for better performance
                data_only=True,   # Get calculated values, not formulas
                keep_vba=False    # Don't load VBA code
            )
            
            self.stats.tabs_processed = len(workbook.worksheets)
            self.logger.debug(f"Workbook loaded successfully with {self.stats.tabs_processed} tabs")
            
            return workbook
            
        except Exception as e:
            self.logger.error(f"Failed to load workbook: {file_path}, error: {str(e)}")
            return None
    
    def _validate_schemas(self, schemas: Dict[str, Dict[str, Any]], 
                         workbook: Workbook) -> List[ValidationResult]:
        """
        Validate schemas against the workbook.
        
        Args:
            schemas: Dictionary of schemas by tab name
            workbook: OpenPyXL workbook object
            
        Returns:
            List of validation results
        """
        self.logger.debug("Validating schemas")
        
        validation_results = []
        
        for tab_name, schema in schemas.items():
            try:
                # Validate schema definition
                schema_result = self.schema_validator.validate_schema_definition(schema)
                validation_results.append(schema_result)
                
                if schema_result.is_valid:
                    # Validate cell references
                    cell_result = self.schema_validator.validate_cell_references(schema, workbook)
                    validation_results.append(cell_result)
                    
                    # Validate data types
                    type_result = self.schema_validator.validate_data_types(schema)
                    validation_results.append(type_result)
                    
                    # Validate field definitions
                    field_result = self.schema_validator.validate_field_definitions(schema)
                    validation_results.append(field_result)
                
            except Exception as e:
                self.logger.error(f"Schema validation error for tab '{tab_name}': {str(e)}")
                validation_results.append(ValidationResult(
                    field_name="schema_validation",
                    is_valid=False,
                    message=f"Schema validation error: {str(e)}",
                    severity="ERROR",
                    location=f"schema.{tab_name}"
                ))
        
        return validation_results
    
    def _extract_data(self, workbook: Workbook, 
                     schemas: Optional[Dict[str, Dict[str, Any]]]) -> Dict[str, Any]:
        """
        Extract data from the workbook.
        
        Args:
            workbook: OpenPyXL workbook object
            schemas: Optional dictionary of schemas by tab name
            
        Returns:
            Dictionary containing extracted data and metadata
        """
        self.logger.debug("Extracting data from workbook")
        
        extraction_results = {
            "tabs": {},
            "metadata": {},
            "processing_stats": {}
        }
        
        try:
            # Extract data from each tab
            for tab_name in workbook.sheetnames:
                self.logger.debug(f"Processing tab: {tab_name}")
                
                schema = schemas.get(tab_name) if schemas else None
                tab_data = self.process_tab(workbook, tab_name, schema)
                
                extraction_results["tabs"][tab_name] = tab_data
                self.stats.fields_processed += len(tab_data.get("data", {}))
            
            # Add metadata
            extraction_results["metadata"] = {
                "total_tabs": len(workbook.sheetnames),
                "tab_names": list(workbook.sheetnames),
                "workbook_properties": self._extract_workbook_properties(workbook)
            }
            
            # Add processing statistics
            extraction_results["processing_stats"] = self.stats.to_dict()
            
        except Exception as e:
            self.logger.error(f"Data extraction error: {str(e)}")
            extraction_results["error"] = str(e)
        
        return extraction_results
    
    def _extract_tab_with_schema(self, worksheet, schema: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract data from a tab using a specific schema.
        
        Args:
            worksheet: OpenPyXL worksheet object
            schema: Schema definition for the tab
            
        Returns:
            Dictionary containing extracted data and validation results
        """
        try:
            fields = schema.get("fields", [])
            extracted_data = {}
            field_validation_results = []
            
            for field in fields:
                field_name = field.get("name")
                cell_ref = field.get("cell_reference")
                data_type = field.get("data_type")
                
                if not all([field_name, cell_ref, data_type]):
                    continue
                
                # Extract cell value
                cell_value = self._get_cell_value(worksheet, cell_ref)
                
                # Validate cell value
                validation_result = self.data_validator.validate_cell_value(
                    cell_value, self._get_python_type(data_type), field_name
                )
                field_validation_results.append(validation_result)
                
                # Store extracted value
                if validation_result.is_valid:
                    # Get the processed value from context if available
                    context = validation_result.context
                    if "converted_value" in context:
                        extracted_data[field_name] = context["converted_value"]
                    else:
                        extracted_data[field_name] = cell_value
                else:
                    # Store original value even if validation failed
                    extracted_data[field_name] = cell_value
                
                self.stats.cells_processed += 1
            
            return {
                "data": extracted_data,
                "validation_results": field_validation_results,
                "schema": schema.get("schema_name", "unknown")
            }
            
        except Exception as e:
            self.logger.error(f"Schema-based extraction error: {str(e)}")
            return {
                "data": {},
                "validation_results": [],
                "error": str(e)
            }
    
    def _extract_tab_default(self, worksheet) -> Dict[str, Any]:
        """
        Extract data from a tab using default extraction logic.
        
        Args:
            worksheet: OpenPyXL worksheet object
            
        Returns:
            Dictionary containing extracted data
        """
        try:
            # Get the used range
            min_row = worksheet.min_row or 1
            max_row = worksheet.max_row or 1
            min_col = worksheet.min_column or 1
            max_col = worksheet.max_column or 1
            
            extracted_data = {}
            
            # Extract data from used range
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_ref = cell.coordinate
                        extracted_data[cell_ref] = cell.value
                        self.stats.cells_processed += 1
            
            return {
                "data": extracted_data,
                "range": {
                    "min_row": min_row,
                    "max_row": max_row,
                    "min_col": min_col,
                    "max_col": max_col
                }
            }
            
        except Exception as e:
            self.logger.error(f"Default extraction error: {str(e)}")
            return {
                "data": {},
                "error": str(e)
            }
    
    def _get_cell_value(self, worksheet, cell_ref: str) -> Any:
        """
        Get the value from a specific cell.
        
        Args:
            worksheet: OpenPyXL worksheet object
            cell_ref: Cell reference (e.g., "B15")
            
        Returns:
            Cell value
        """
        try:
            return worksheet[cell_ref].value
        except Exception:
            return None
    
    def _get_python_type(self, data_type: str) -> type:
        """
        Convert schema data type to Python type.
        
        Args:
            data_type: Schema data type string
            
        Returns:
            Python type
        """
        type_mapping = {
            'string': str,
            'str': str,
            'text': str,
            'integer': int,
            'int': int,
            'number': int,
            'float': float,
            'decimal': float,
            'double': float,
            'boolean': bool,
            'bool': bool,
            'logical': bool,
            'datetime': str,  # Keep as string for now
            'date': str,
            'time': str,
            'email': str,
            'url': str,
            'hyperlink': str
        }
        
        normalized_type = data_type.lower().strip()
        return type_mapping.get(normalized_type, str)
    
    def _extract_workbook_properties(self, workbook: Workbook) -> Dict[str, Any]:
        """
        Extract basic workbook properties.
        
        Args:
            workbook: OpenPyXL workbook object
            
        Returns:
            Dictionary of workbook properties
        """
        try:
            props = workbook.properties
            return {
                "title": getattr(props, 'title', None),
                "creator": getattr(props, 'creator', None),
                "created": getattr(props, 'created', None),
                "modified": getattr(props, 'modified', None),
                "last_modified_by": getattr(props, 'lastModifiedBy', None)
            }
        except Exception:
            return {}
    
    def _create_success_result(self, file_path: Path, 
                              validation_results: List[ValidationResult],
                              extraction_results: Dict[str, Any],
                              start_time: float) -> ExcelParseResult:
        """
        Create a successful processing result.
        
        Args:
            file_path: Path to the processed file
            validation_results: List of validation results
            extraction_results: Results of data extraction
            start_time: Start time of processing
            
        Returns:
            ExcelParseResult object
        """
        end_time = time.time()
        processing_time = end_time - start_time
        
        # Update statistics
        self.stats.end_time = end_time
        self.stats.validation_errors = len([r for r in validation_results if not r.is_valid])
        
        # Determine success based on validation results
        critical_errors = [
            r for r in validation_results 
            if not r.is_valid and r.severity == "ERROR"
        ]
        
        success = len(critical_errors) == 0 or not self.config.strict_mode
        
        return ExcelParseResult(
            success=success,
            metadata=extraction_results.get("metadata", {}),
            schemas={},  # Will be populated if schemas were provided
            tab_contents=extraction_results.get("tabs", {}),
            validation_results=validation_results,
            processing_stats=self.stats,
            errors=[r.message for r in validation_results if not r.is_valid and r.severity == "ERROR"],
            warnings=[r.message for r in validation_results if not r.is_valid and r.severity == "WARNING"],
            file_path=file_path,
            processing_time=processing_time
        )
    
    def _create_failed_result(self, file_path: Path,
                             validation_results: List[ValidationResult],
                             start_time: float,
                             error_message: str) -> ExcelParseResult:
        """
        Create a failed processing result.
        
        Args:
            file_path: Path to the processed file
            validation_results: List of validation results
            start_time: Start time of processing
            error_message: Error message describing the failure
            
        Returns:
            ExcelParseResult object
        """
        end_time = time.time()
        processing_time = end_time - start_time
        
        # Update statistics
        self.stats.end_time = end_time
        self.stats.validation_errors = len([r for r in validation_results if not r.is_valid])
        
        return ExcelParseResult(
            success=False,
            metadata={},
            schemas={},
            tab_contents={},
            validation_results=validation_results,
            processing_stats=self.stats,
            errors=[error_message] + [r.message for r in validation_results if not r.is_valid and r.severity == "ERROR"],
            warnings=[r.message for r in validation_results if not r.is_valid and r.severity == "WARNING"],
            file_path=file_path,
            processing_time=processing_time
        )
