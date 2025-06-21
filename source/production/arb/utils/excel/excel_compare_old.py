# === START OF FILE ===
"""
excel_compare.py

Compare Excel workbooks (.xlsx) individually or across two directories.

Features:
  - File metadata (created, modified)
  - SHA-256 binary comparison
  - Cell-by-cell value comparison
  - Formula detection and comparison (including array formulas)
  - Comment comparison
  - Optional formatting comparison:
      - "off": skip formatting
      - "common": compare bold, italic, font size, fill, alignment
      - "full": compare all formatting including borders, protection
  - Sheet mismatch detection:
      - If sheet names differ, lists sheets only in A and only in B
      - Still compares common sheets

Output:
  Results are written to a timestamped text file:
    comparison_at_YYYYMMDD_HHMMSS.txt

Example usage:
    compare_excel_directories(
      Path("C:/local/folder"),
      Path("C:/onedrive/folder"),
      formatting_mode="full"
    )
"""

import hashlib
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from typing import Union

COMMON_FONT_ATTRS = ["bold", "italic", "size", "name"]
FULL_FONT_ATTRS = [
  "bold", "italic", "underline", "strike", "size", "name", "color", "vertAlign",
  "charset", "scheme", "family", "outline", "shadow"
]
FILL_ATTRS = ["patternType", "fgColor", "bgColor"]
ALIGNMENT_ATTRS = [
  "horizontal", "vertical", "wrapText", "shrinkToFit", "indent",
  "readingOrder", "textRotation"
]
PROTECTION_ATTRS = ["locked", "hidden"]
BORDER_SIDES = ["left", "right", "top", "bottom", "diagonal"]

def stringify_formula(cell) -> str:
  """
  Return a human-readable formula string or raw value for a cell.

  Args:
    cell (Cell): openpyxl cell object.

  Returns:
    str: Formula string, or raw cell value if not a formula.
  """
  try:
    if isinstance(cell.value, ArrayFormula):
      return cell.value.formula
    if cell.data_type == 'f':
      return str(cell.formula or cell._value or cell.value)
    return str(cell.value)
  except Exception as e:
    return f"[unreadable: {e}]"

def compute_sha256(path: Path) -> str:
  with open(path, "rb") as f:
    return hashlib.sha256(f.read()).hexdigest()

def compare_excel_content(path1: Path, path2: Path, formatting_mode: str = "common") -> list[str]:
  wb1 = load_workbook(path1, data_only=False)
  wb2 = load_workbook(path2, data_only=False)
  differences = []

  sheets1 = set(wb1.sheetnames)
  sheets2 = set(wb2.sheetnames)

  only_in_a = sorted(sheets1 - sheets2)
  only_in_b = sorted(sheets2 - sheets1)
  in_both = sorted(sheets1 & sheets2)

  if only_in_a:
    differences.append(f"Sheets only in A: {only_in_a}")
  if only_in_b:
    differences.append(f"Sheets only in B: {only_in_b}")

  for sheet_name in in_both:
    ws1 = wb1[sheet_name]
    ws2 = wb2[sheet_name]
    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    for row in range(1, max_row + 1):
      for col in range(1, max_col + 1):
        cell1 = ws1.cell(row=row, column=col)
        cell2 = ws2.cell(row=row, column=col)
        coord = cell1.coordinate

        val1 = stringify_formula(cell1)
        val2 = stringify_formula(cell2)

        if val1 != val2:
          differences.append(f"{sheet_name} {coord}: '{val1}' != '{val2}'")

        if cell1.data_type == 'f' and cell2.data_type == 'f':
          if val1 != val2:
            differences.append(f"{sheet_name} {coord}: formulas differ '{val1}' vs '{val2}'")
        elif cell1.data_type == 'f' or cell2.data_type == 'f':
          differences.append(f"{sheet_name} {coord}: formula presence mismatch")

        comment1 = cell1.comment.text if cell1.comment else None
        comment2 = cell2.comment.text if cell2.comment else None
        if comment1 != comment2:
          differences.append(f"{sheet_name} {coord}: comment changed\n  A: {comment1}\n  B: {comment2}")

        if formatting_mode != "off":
          font1, font2 = cell1.font, cell2.font
          font_attrs = COMMON_FONT_ATTRS if formatting_mode == "common" else FULL_FONT_ATTRS
          for attr in font_attrs:
            if getattr(font1, attr, None) != getattr(font2, attr, None):
              differences.append(f"{sheet_name} {coord}: font.{attr} changed")

          fill1, fill2 = cell1.fill, cell2.fill
          for attr in FILL_ATTRS:
            if getattr(fill1, attr, None) != getattr(fill2, attr, None):
              differences.append(f"{sheet_name} {coord}: fill.{attr} changed")

          align1, align2 = cell1.alignment, cell2.alignment
          for attr in ALIGNMENT_ATTRS:
            if getattr(align1, attr, None) != getattr(align2, attr, None):
              differences.append(f"{sheet_name} {coord}: alignment.{attr} changed")

          if formatting_mode == "full":
            border1, border2 = cell1.border, cell2.border
            for side in BORDER_SIDES:
              b1 = getattr(border1, side)
              b2 = getattr(border2, side)
              if b1.style != b2.style or b1.color != b2.color:
                differences.append(f"{sheet_name} {coord}: border.{side} changed")

            prot1, prot2 = cell1.protection, cell2.protection
            for attr in PROTECTION_ATTRS:
              if getattr(prot1, attr, None) != getattr(prot2, attr, None):
                differences.append(f"{sheet_name} {coord}: protection.{attr} changed")

  return differences

def compare_excel_files(file_a_path: Union[str, Path], file_b_path: Union[str, Path],
                        formatting_mode: str = "common") -> list[str]:
  file_a = Path(file_a_path)
  file_b = Path(file_b_path)
  output = []

  output.append(f"Comparing:\n  A: {file_a.name}\n  B: {file_b.name}\n")

  hash_a = compute_sha256(file_a)
  hash_b = compute_sha256(file_b)

  output.append("SHA-256 Hashes:")
  output.append(f"  A: {hash_a}")
  output.append(f"  B: {hash_b}")

  if hash_a == hash_b:
    output.append("✔ Files are identical at the binary level.\n")
    return output

  output.append("→ Hashes differ; comparing cell content...\n")
  diffs = compare_excel_content(file_a, file_b, formatting_mode=formatting_mode)

  if not diffs:
    output.append("✔ No cell/comment/formatting differences found.\n")
  else:
    output.extend(diffs)

  return output

def compare_excel_directories(dir_a: Path, dir_b: Path, formatting_mode: str = "common") -> None:
  now = datetime.now()
  out_path = Path(f"comparison_at_{now.strftime('%Y%m%d_%H%M%S')}.txt")

  excel_files_a = {f.name: f for f in dir_a.glob("*.xlsx") if f.is_file()}
  excel_files_b = {f.name: f for f in dir_b.glob("*.xlsx") if f.is_file()}

  only_in_a = sorted(set(excel_files_a) - set(excel_files_b))
  only_in_b = sorted(set(excel_files_b) - set(excel_files_a))
  in_both = sorted(set(excel_files_a) & set(excel_files_b))

  with open(out_path, "w", encoding="utf-8") as f:
    f.write(f"Comparing Excel files in:\n  A: {dir_a}\n  B: {dir_b}\n\n")
    f.write(f"Files only in A ({len(only_in_a)}):\n")
    for name in only_in_a:
      f.write(f"  {name}\n")
    f.write(f"\nFiles only in B ({len(only_in_b)}):\n")
    for name in only_in_b:
      f.write(f"  {name}\n")

    f.write(f"\nFiles in both ({len(in_both)}):\n")
    for name in in_both:
      f.write(f"\n=== Comparing {name} ===\n")
      result = compare_excel_files(
        excel_files_a[name],
        excel_files_b[name],
        formatting_mode=formatting_mode
      )
      for line in result:
        f.write(line + "\n")

  print(f"Comparison complete. Output saved to: {out_path.resolve()}")


if __name__ == "__main__":
  local_path = Path(r"C:\tony_local\pycharm\feedback_portal\feedback_forms\processed_versions\xl_workbooks")
  sharepoint_path = Path(r"C:\one_drive\OneDriveLinks\RD Satellite Project - Operator Notification Materials for Review\spreadsheets\xl_workbooks")

  compare_excel_directories(
    local_path,
    sharepoint_path,
    formatting_mode="common"
  )
