import warnings
from datetime import datetime
from os import path

import openpyxl

TEMPLATE_LOCATION = "C:/output/webapp_resources/"


def fill_spreadsheet(template='industrial',
                     edit=True,
                     id_incidence=10001,
                     coordinates='42.1111, -119.2222',
                     plume_timestamp="2025-01-02 10:00:00"):
  """
  test filling & saving operator spreadsheet templates
  """
  # open & fill Excel template for attachment:
  if template == 'industrial':
    template_file = path.join(TEMPLATE_LOCATION, "og_operator_feedback_with_formatting_v11.xlsx")
  else:
    template_file = path.join(TEMPLATE_LOCATION, "landfill_operator_feedback_with_formatting_v15.xlsx")

  # suppress warnings about data validation - Excel's extensions aren't fully compatible with openpyxl but these modified templates should work
  # https://www.reddit.com/r/learnpython/comments/mfy9qa/openpyxl_deleting_data_validation/
  with warnings.catch_warnings():
    warnings.filterwarnings("ignore", category=UserWarning)
    wb = openpyxl.load_workbook(template_file, keep_vba=False)
  if edit:
    ws = wb['Feedback Form']
    if template == 'industrial':
      ws["$D$11"] = id_incidence
      ws["$D$12"] = datetime.now().date().strftime("%m/%d/%Y")
      ws["$D$13"] = coordinates
      ws["$D$14"] = plume_timestamp
    else:
      ws["$C$10"] = id_incidence
      ws["$C$11"] = ', '.join([str(id_plume) for id_plume in nonempty_list])
      ws["$C$12"] = coordinates
      ws["$C$13"] = plume_timestamp

  with open(path.join(TEMPLATE_LOCATION, "test_output.xlsx"), 'wb') as f:
    wb.save(f)
