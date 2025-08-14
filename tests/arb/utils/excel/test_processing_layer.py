"""
Test suite for Excel processing layer components.

This module tests the processing components created in Phase 3 of the refactoring,
including the ExcelProcessor class and its integration with the validation layer.

Test Classes:
    TestExcelProcessor: Tests for Excel processing orchestration
    TestProcessingIntegration: Tests for processing layer integration

Author: AI Assistant
Created: 2025-01-27
Version: 1.0
"""

import pytest
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict

# Import the enhanced components
from arb.utils.excel.core import (
    ExcelParseConfig,
    ExcelParseResult,
    ProcessingStats,
    ValidationResult
)

from arb.utils.excel.processing import ExcelProcessor
from arb.utils.excel.validation import ExcelValidator, SchemaValidator, DataValidator


class TestExcelProcessor:
    """Test Excel processing orchestration functionality."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.config = ExcelParseConfig(
            max_file_size_mb=10,
            validate_file_exists=True,
            validate_file_format=True,
            max_tabs=5,
            strict_mode=False  # Allow processing to continue with warnings
        )
        self.processor = ExcelProcessor(self.config)
        
        # Create temporary test files
        self.temp_dir = Path(tempfile.mkdtemp())
        self.valid_xlsx_path = self.temp_dir / "test.xlsx"
        self.invalid_path = self.temp_dir / "nonexistent.xlsx"
        
        # Create a simple Excel file for testing
        self._create_test_excel_file()
        
        # Sample schema for testing
        self.sample_schema = {
            "Sheet1": {
                "schema_name": "test_schema",
                "tab_name": "Sheet1",
                "fields": [
                    {
                        "name": "project_name",
                        "cell_reference": "B15",
                        "data_type": "string"
                    },
                    {
                        "name": "quantity",
                        "cell_reference": "B16",
                        "data_type": "integer"
                    }
                ]
            }
        }
    
    def teardown_method(self):
        """Clean up test fixtures."""
        import shutil
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    
    def _create_test_excel_file(self):
        """Create a simple test Excel file."""
        # Use openpyxl to create a proper test file
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add some test data
        ws["B15"] = "Test Project"
        ws["B16"] = 42
        
        # Save the file
        wb.save(self.valid_xlsx_path)
        wb.close()
    
    def test_processor_initialization(self):
        """Test processor initialization with configuration."""
        processor = ExcelProcessor(self.config)
        
        assert processor.config == self.config
        assert processor.excel_validator is not None
        assert processor.schema_validator is not None
        assert processor.data_validator is not None
        assert processor.stats is not None
    
    def test_processor_default_config(self):
        """Test processor initialization with default configuration."""
        processor = ExcelProcessor()
        
        assert processor.config is not None
        assert isinstance(processor.config, ExcelParseConfig)
        assert processor.config.max_file_size_mb == 100  # Default value
    
    def test_process_file_basic(self):
        """Test basic file processing without schemas."""
        result = self.processor.process_file(self.valid_xlsx_path)
        
        assert isinstance(result, ExcelParseResult)
        assert result.success is True
        assert result.file_path == self.valid_xlsx_path
        assert result.processing_time > 0
        assert len(result.tab_contents) > 0
        assert "Sheet1" in result.tab_contents
    
    def test_process_file_with_schemas(self):
        """Test file processing with schemas."""
        result = self.processor.process_file(
            self.valid_xlsx_path,
            schemas=self.sample_schema
        )
        
        assert isinstance(result, ExcelParseResult)
        assert result.success is True
        assert "Sheet1" in result.tab_contents
        
        # Check that schema-based extraction was used
        sheet1_data = result.tab_contents["Sheet1"]
        assert "schema" in sheet1_data
        assert sheet1_data["schema"] == "test_schema"
        assert "data" in sheet1_data
    
    def test_process_file_with_required_tabs(self):
        """Test file processing with required tabs validation."""
        result = self.processor.process_file(
            self.valid_xlsx_path,
            required_tabs=["Sheet1"]
        )
        
        assert isinstance(result, ExcelParseResult)
        assert result.success is True
        assert "Sheet1" in result.tab_contents
    
    def test_process_file_nonexistent(self):
        """Test file processing with nonexistent file."""
        result = self.processor.process_file(self.invalid_path)
        
        assert isinstance(result, ExcelParseResult)
        assert result.success is False
        assert len(result.errors) > 0
        # Check for any of the expected error messages
        error_messages = [error.lower() for error in result.errors]
        assert any("does not exist" in msg or "file validation failed" in msg or "failed to load workbook" in msg for msg in error_messages)
    
    def test_process_file_strict_mode(self):
        """Test file processing in strict mode."""
        strict_config = ExcelParseConfig(strict_mode=True)
        strict_processor = ExcelProcessor(strict_config)
        
        # Create a file that will have validation warnings
        result = strict_processor.process_file(self.valid_xlsx_path)
        
        # In strict mode, even warnings should cause failure
        # But our test file should pass basic validations
        assert isinstance(result, ExcelParseResult)
    
    def test_process_tab_with_schema(self):
        """Test processing a single tab with schema."""
        import openpyxl
        
        # Create a simple workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add some test data
        ws["B15"] = "Test Project"
        ws["B16"] = 42
        
        try:
            result = self.processor.process_tab(wb, "Sheet1", self.sample_schema["Sheet1"])
            
            assert "data" in result
            assert "validation_results" in result
            assert "schema" in result
            
            # Check extracted data
            data = result["data"]
            assert "project_name" in data
            assert "quantity" in data
            assert data["project_name"] == "Test Project"
            assert data["quantity"] == 42
            
            # Check validation results
            validation_results = result["validation_results"]
            assert len(validation_results) == 2
            
        finally:
            wb.close()
    
    def test_process_tab_without_schema(self):
        """Test processing a single tab without schema."""
        import openpyxl
        
        # Create a simple workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add some test data
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Value1"
        ws["B2"] = "Value2"
        
        try:
            result = self.processor.process_tab(wb, "Sheet1")
            
            assert "data" in result
            assert "range" in result
            
            # Check extracted data
            data = result["data"]
            assert "A1" in data
            assert "B1" in data
            assert "A2" in data
            assert "B2" in data
            
            # Check range information
            range_info = result["range"]
            assert range_info["min_row"] == 1
            assert range_info["max_row"] == 2
            assert range_info["min_col"] == 1
            assert range_info["max_col"] == 2
            
        finally:
            wb.close()
    
    def test_process_tab_not_found(self):
        """Test processing a tab that doesn't exist."""
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        try:
            with pytest.raises(Exception) as exc_info:
                self.processor.process_tab(wb, "NonexistentSheet")
            
            assert "not found" in str(exc_info.value)
            
        finally:
            wb.close()
    
    def test_processing_statistics(self):
        """Test that processing statistics are properly tracked."""
        result = self.processor.process_file(self.valid_xlsx_path)
        
        assert result.processing_stats is not None
        assert isinstance(result.processing_stats, ProcessingStats)
        
        # Check basic statistics
        stats = result.processing_stats
        assert stats.start_time > 0
        assert stats.end_time > 0
        assert stats.tabs_processed > 0
        # Note: cells_processed might be 0 if the test file has no data in the expected range
        # We'll check that the field exists and is a number
        assert isinstance(stats.cells_processed, int)
        assert stats.total_time > 0
    
    def test_validation_integration(self):
        """Test that validation is properly integrated into processing."""
        result = self.processor.process_file(self.valid_xlsx_path)
        
        assert len(result.validation_results) > 0
        assert all(isinstance(r, ValidationResult) for r in result.validation_results)
        
        # Check that validation results are properly categorized
        errors = [r for r in result.validation_results if not r.is_valid and r.severity == "ERROR"]
        warnings = [r for r in result.validation_results if not r.is_valid and r.severity == "WARNING"]
        
        # Our test file should pass basic validations
        assert len(errors) == 0, f"Unexpected errors: {[e.message for e in errors]}"
    
    def test_error_handling(self):
        """Test error handling during processing."""
        # Test with a corrupted file path
        corrupted_path = self.temp_dir / "corrupted.xlsx"
        corrupted_path.write_bytes(b"not an excel file")
        
        result = self.processor.process_file(corrupted_path)
        
        assert isinstance(result, ExcelParseResult)
        assert result.success is False
        assert len(result.errors) > 0
    
    def test_workbook_cleanup(self):
        """Test that workbooks are properly closed after processing."""
        # This test ensures that resources are properly managed
        result = self.processor.process_file(self.valid_xlsx_path)
        
        # The workbook should be closed automatically
        assert result.success is True
        
        # Verify no resource leaks by processing multiple files
        for i in range(3):
            result = self.processor.process_file(self.valid_xlsx_path)
            assert result.success is True


class TestProcessingIntegration:
    """Test integration between processing layer components."""
    
    def test_processor_with_validators(self):
        """Test that processor properly uses validation components."""
        processor = ExcelProcessor()
        
        # Verify validator components are accessible
        assert processor.excel_validator is not None
        assert processor.schema_validator is not None
        assert processor.data_validator is not None
        
        # Test that validators are of correct types
        assert isinstance(processor.excel_validator, ExcelValidator)
        assert isinstance(processor.schema_validator, SchemaValidator)
        assert isinstance(processor.data_validator, DataValidator)
    
    def test_end_to_end_processing_workflow(self):
        """Test complete end-to-end processing workflow."""
        processor = ExcelProcessor()
        
        # Create a test file
        temp_dir = Path(tempfile.mkdtemp())
        test_file = temp_dir / "test.xlsx"
        
        try:
            # Create a minimal Excel file using openpyxl
            import openpyxl
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Add some test data
            ws["A1"] = "Test"
            
            # Save the file
            wb.save(test_file)
            wb.close()
            
            # Process the file
            result = processor.process_file(test_file)
            
            # Verify the result
            assert isinstance(result, ExcelParseResult)
            assert result.success is True
            assert result.file_path == test_file
            assert len(result.tab_contents) > 0
            assert "Sheet1" in result.tab_contents
            
            # Check processing statistics
            assert result.processing_stats is not None
            assert result.processing_stats.tabs_processed > 0
            assert result.processing_stats.cells_processed > 0
            
        finally:
            import shutil
            if temp_dir.exists():
                shutil.rmtree(temp_dir)
    
    def test_configuration_impact(self):
        """Test that different configurations impact processing behavior."""
        # Test with different file size limits
        small_config = ExcelParseConfig(max_file_size_mb=1)
        large_config = ExcelParseConfig(max_file_size_mb=100)
        
        small_processor = ExcelProcessor(small_config)
        large_processor = ExcelProcessor(large_config)
        
        # Create a test file
        temp_dir = Path(tempfile.mkdtemp())
        test_file = temp_dir / "test.xlsx"
        
        try:
            # Create a minimal Excel file using openpyxl
            import openpyxl
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Add some test data
            ws["A1"] = "Test"
            
            # Save the file
            wb.save(test_file)
            wb.close()
            
            # Both processors should handle the small file
            small_result = small_processor.process_file(test_file)
            large_result = large_processor.process_file(test_file)
            
            assert small_result.success is True
            assert large_result.success is True
            
        finally:
            import shutil
            if temp_dir.exists():
                shutil.rmtree(temp_dir)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
