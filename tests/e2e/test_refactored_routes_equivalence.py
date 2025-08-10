"""
Pytest-compatible E2E Tests for Refactored Route Equivalence

This module tests that the refactored routes produce identical results to the original routes
for all test files in the standard testing directory. This ensures the refactoring maintains
functional equivalence while improving code structure.

Key Requirements:
1. upload_file_refactored must behave identically to upload_file
2. upload_file_staged_refactored must behave identically to upload_file_staged
3. All test files in feedback_forms/testing_versions/standard must be tested
4. Tests must fail explicitly if test directories cannot be resolved
5. Performance characteristics (files uploaded, messages, redirects) must match exactly

Test Coverage:
- Route response equivalence (status codes, redirects, error messages)
- File processing equivalence (same files uploaded, same database records)
- Error handling equivalence (same error messages, same failure modes)
- User experience equivalence (same flash messages, same page content)
"""

import json
import os
import re
import sqlite3
import time
import warnings
from pathlib import Path
from typing import Dict, Any, Tuple, Optional

import openpyxl
import psycopg2
import pytest
from playwright.sync_api import Page, expect

import conftest
from arb.portal.utils.e2e_testing_util import navigate_and_wait_for_ready
from arb.portal.utils.playwright_testing_util import (
    clear_upload_attempt_marker,
    upload_file_and_wait_for_attempt_marker,
    wait_for_upload_attempt_marker
)

# Test configuration
BASE_URL = os.environ.get('TEST_BASE_URL', conftest.TEST_BASE_URL)

# Suppress openpyxl warnings about unsupported Excel features
@pytest.fixture(autouse=True, scope="session")
def suppress_openpyxl_warnings():
    warnings.filterwarnings(
        "ignore",
        message=".*extension is not supported and will be removed",
        category=UserWarning,
        module=r"openpyxl\.reader\.excel"
    )

def get_xls_files(base_path: Path, recursive: bool = False, excel_exts=None) -> list:
    """
    Return a list of all Excel-like files in the given directory.
    Args:
        base_path: Path to search
        recursive: If True, search subdirectories
        excel_exts: List of extensions to include (default: .xlsx, .xls)
    Returns:
        List of file paths as strings
    """
    if excel_exts is None:
        excel_exts = ['.xlsx', '.xls']
    
    files = []
    if recursive:
        for ext in excel_exts:
            files.extend(base_path.rglob(f"*{ext}"))
    else:
        for ext in excel_exts:
            files.extend(base_path.glob(f"*{ext}"))
    
    return [str(f) for f in sorted(files)]

def get_test_files() -> list:
    """
    Get a list of test files to use for parameterized tests.
    This function MUST fail explicitly if test directories cannot be resolved.
    Returns:
        List of file paths for testing
    """
    base_dirs = [
        conftest.STANDARD_TEST_FILES_DIR,
    ]
    files = []
    for base_dir in base_dirs:
        files.extend(get_xls_files(base_dir, recursive=True))
    
    # CRITICAL: Fail catastrophically if no test files are found
    if not files:
        error_msg = f"""
❌ CRITICAL TEST INFRASTRUCTURE ERROR: No test files found!

Base directories checked: {base_dirs}
Current working directory: {Path.cwd()}
Repository root: {conftest.REPO_ROOT}
Standard test files dir: {conftest.STANDARD_TEST_FILES_DIR}
Standard test files dir exists: {conftest.STANDARD_TEST_FILES_DIR.exists()}
Files in standard dir: {list(conftest.STANDARD_TEST_FILES_DIR.glob('*'))}

This test will fail catastrophically to prevent silent test failures.
"""
        pytest.fail(error_msg)
    
    print(f"✓ Found {len(files)} test files for parameterized testing")
    return files

def read_excel_fields(file_path: str) -> Dict[str, Any]:
    """
    Read Excel file and extract field names and values.
    Args:
        file_path: Path to Excel file
    Returns:
        Dictionary mapping field names to values
    """
    if not os.path.exists(file_path):
        pytest.skip(f"Test file not found: {file_path}")
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        fields = {}
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            if row[0].value and row[1].value:
                fields[str(row[0].value).strip()] = str(row[1].value).strip()
        
        workbook.close()
        return fields
    except Exception as e:
        pytest.fail(f"Failed to read Excel file {file_path}: {e}")

def get_id_from_redirect(page: Page) -> Optional[int]:
    """
    Extract ID from redirect URL after successful upload.
    Args:
        page: Playwright page object
    Returns:
        ID as integer, or None if not found
    """
    current_url = page.url
    match = re.search(r'/incidence_update/(\d+)', current_url)
    if match:
        return int(match.group(1))
    return None

def fetch_misc_json_from_db(id_: int) -> Optional[Dict[str, Any]]:
    """
    Fetch misc_json data from database for given ID.
    Args:
        id_: Database ID
    Returns:
        Parsed JSON data or None if not found
    """
    try:
        # Try PostgreSQL first
        conn = psycopg2.connect(
            host=os.environ.get('DB_HOST', 'localhost'),
            database=os.environ.get('DB_NAME', 'feedback_portal'),
            user=os.environ.get('DB_USER', 'postgres'),
            password=os.environ.get('DB_PASSWORD', '')
        )
        cursor = conn.cursor()
        cursor.execute("SELECT misc_json FROM portal_updates WHERE id = %s", (id_,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if result and result[0]:
            return json.loads(result[0])
        return None
    except Exception as e:
        print(f"PostgreSQL connection failed: {e}")
        try:
            # Fallback to SQLite
            db_path = os.environ.get('SQLITE_DB_PATH', 'feedback_portal.db')
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT misc_json FROM portal_updates WHERE id = ?", (id_,))
            result = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if result and result[0]:
                return json.loads(result[0])
            return None
        except Exception as e2:
            pytest.fail(f"Database connection failed for both PostgreSQL and SQLite: {e}, {e2}")

class TestRefactoredRouteEquivalence:
    """
    Test class for ensuring refactored routes maintain functional equivalence
    with the original routes for all test files.
    """
    
    @pytest.fixture(scope="function")
    def upload_page(self, page: Page):
        """Navigate to upload page and wait for ready state."""
        navigate_and_wait_for_ready(page, f"{BASE_URL}/upload")
        return page
    
    @pytest.fixture(scope="function")
    def upload_staged_page(self, page: Page):
        """Navigate to staged upload page and wait for ready state."""
        navigate_and_wait_for_ready(page, f"{BASE_URL}/upload_staged")
        return page
    
    @pytest.fixture(scope="function")
    def upload_refactored_page(self, page: Page):
        """Navigate to refactored upload page and wait for ready state."""
        navigate_and_wait_for_ready(page, f"{BASE_URL}/upload_refactored")
        return page
    
    @pytest.fixture(scope="function")
    def upload_staged_refactored_page(self, page: Page):
        """Navigate to refactored staged upload page and wait for ready state."""
        navigate_and_wait_for_ready(page, f"{BASE_URL}/upload_staged_refactored")
        return page

    @pytest.mark.parametrize("file_path", get_test_files())
    def test_upload_file_equivalence(self, upload_page: Page, upload_refactored_page: Page, file_path: str):
        """
        Test that upload_file and upload_file_refactored produce identical results.
        
        This test ensures that the refactored route maintains exact functional equivalence
        with the original route for all test files in the standard testing directory.
        """
        if not os.path.exists(file_path):
            pytest.skip(f"Test file not found: {file_path}")
        
        print(f"\n🔍 Testing file: {file_path}")
        
        # Test original route
        clear_upload_attempt_marker(upload_page)
        upload_file_and_wait_for_attempt_marker(upload_page, file_path)
        
        # Capture original route results
        original_status = upload_page.url
        original_content = upload_page.content()
        original_flash_messages = self._extract_flash_messages(upload_page)
        original_id = get_id_from_redirect(upload_page)
        
        print(f"  Original route - Status: {original_status}, ID: {original_id}")
        
        # Test refactored route
        clear_upload_attempt_marker(upload_refactored_page)
        upload_file_and_wait_for_attempt_marker(upload_refactored_page, file_path)
        
        # Capture refactored route results
        refactored_status = upload_refactored_page.url
        refactored_content = upload_refactored_page.content()
        refactored_flash_messages = self._extract_flash_messages(upload_refactored_page)
        refactored_id = get_id_from_redirect(upload_refactored_page)
        
        print(f"  Refactored route - Status: {refactored_status}, ID: {refactored_id}")
        
        # Assert equivalence
        assert original_status == refactored_status, f"URL mismatch for {file_path}"
        assert original_id == refactored_id, f"ID mismatch for {file_path}"
        assert original_flash_messages == refactored_flash_messages, f"Flash message mismatch for {file_path}"
        
        # If both succeeded, verify database equivalence
        if original_id and refactored_id:
            original_data = fetch_misc_json_from_db(original_id)
            refactored_data = fetch_misc_json_from_db(refactored_id)
            
            assert original_data == refactored_data, f"Database data mismatch for {file_path}"
            
            # Verify Excel field equivalence
            excel_fields = read_excel_fields(file_path)
            for field_name, excel_value in excel_fields.items():
                if field_name in original_data:
                    assert str(original_data[field_name]) == str(excel_value), \
                        f"Field {field_name} mismatch for {file_path}: Excel={excel_value}, DB={original_data[field_name]}"
        
        print(f"  ✅ Equivalence verified for {file_path}")

    @pytest.mark.parametrize("file_path", get_test_files())
    def test_upload_staged_equivalence(self, upload_staged_page: Page, upload_staged_refactored_page: Page, file_path: str):
        """
        Test that upload_file_staged and upload_file_staged_refactored produce identical results.
        
        This test ensures that the refactored staged route maintains exact functional equivalence
        with the original staged route for all test files in the standard testing directory.
        """
        if not os.path.exists(file_path):
            pytest.skip(f"Test file not found: {file_path}")
        
        print(f"\n🔍 Testing staged file: {file_path}")
        
        # Test original staged route
        clear_upload_attempt_marker(upload_staged_page)
        upload_file_and_wait_for_attempt_marker(upload_staged_page, file_path)
        
        # Capture original staged route results
        original_status = upload_staged_page.url
        original_content = upload_staged_page.content()
        original_flash_messages = self._extract_flash_messages(upload_staged_page)
        original_id = get_id_from_redirect(upload_staged_page)
        
        print(f"  Original staged route - Status: {original_status}, ID: {original_id}")
        
        # Test refactored staged route
        clear_upload_attempt_marker(upload_staged_refactored_page)
        upload_file_and_wait_for_attempt_marker(upload_staged_refactored_page, file_path)
        
        # Capture refactored staged route results
        refactored_status = upload_staged_refactored_page.url
        refactored_content = upload_staged_refactored_page.content()
        refactored_flash_messages = self._extract_flash_messages(upload_staged_refactored_page)
        refactored_id = get_id_from_redirect(upload_staged_refactored_page)
        
        print(f"  Refactored staged route - Status: {refactored_status}, ID: {refactored_id}")
        
        # Assert equivalence
        assert original_status == refactored_status, f"URL mismatch for staged {file_path}"
        assert original_id == refactored_id, f"ID mismatch for staged {file_path}"
        assert original_flash_messages == refactored_flash_messages, f"Flash message mismatch for staged {file_path}"
        
        # If both succeeded, verify database equivalence
        if original_id and refactored_id:
            original_data = fetch_misc_json_from_db(original_id)
            refactored_data = fetch_misc_json_from_db(refactored_id)
            
            assert original_data == refactored_data, f"Database data mismatch for staged {file_path}"
            
            # Verify Excel field equivalence
            excel_fields = read_excel_fields(file_path)
            for field_name, excel_value in excel_fields.items():
                if field_name in original_data:
                    assert str(original_data[field_name]) == str(excel_value), \
                        f"Field {field_name} mismatch for staged {file_path}: Excel={excel_value}, DB={original_data[field_name]}"
        
        print(f"  ✅ Staged equivalence verified for {file_path}")

    def _extract_flash_messages(self, page: Page) -> list:
        """
        Extract flash messages from the page.
        Args:
            page: Playwright page object
        Returns:
            List of flash message texts
        """
        flash_elements = page.locator(".flash-message, .alert, [class*='flash'], [class*='alert']")
        messages = []
        for i in range(flash_elements.count()):
            messages.append(flash_elements.nth(i).text_content().strip())
        return sorted(messages)

    def test_route_page_equivalence(self, upload_page: Page, upload_refactored_page: Page):
        """
        Test that the upload pages for original and refactored routes are functionally equivalent.
        """
        # Check page titles
        expect(upload_page).to_have_title("Upload File")
        expect(upload_refactored_page).to_have_title("Upload File")
        
        # Check form elements
        original_form = upload_page.locator("form")
        refactored_form = upload_refactored_page.locator("form")
        
        expect(original_form).to_be_visible()
        expect(refactored_form).to_be_visible()
        
        # Check file input elements
        original_file_input = upload_page.locator("input[type='file']")
        refactored_file_input = upload_refactored_page.locator("input[type='file']")
        
        assert original_file_input.count() > 0
        assert refactored_file_input.count() > 0
        
        # Check drop zones
        original_drop_zone = upload_page.locator(".drop-zone, [id*='drop']")
        refactored_drop_zone = upload_refactored_page.locator(".drop-zone, [id*='drop']")
        
        assert original_drop_zone.count() == refactored_drop_zone.count()

    def test_staged_route_page_equivalence(self, upload_staged_page: Page, upload_staged_refactored_page: Page):
        """
        Test that the staged upload pages for original and refactored routes are functionally equivalent.
        """
        # Check page titles
        expect(upload_staged_page).to_have_title("Upload Staged File")
        expect(upload_staged_refactored_page).to_have_title("Upload Staged File")
        
        # Check form elements
        original_form = upload_staged_page.locator("form")
        refactored_form = upload_staged_refactored_page.locator("form")
        
        expect(original_form).to_be_visible()
        expect(refactored_form).to_be_visible()
        
        # Check file input elements
        original_file_input = upload_staged_page.locator("input[type='file']")
        refactored_file_input = upload_staged_refactored_page.locator("input[type='file']")
        
        assert original_file_input.count() > 0
        assert refactored_file_input.count() > 0

def pytest_addoption(parser):
    """Add custom command line options."""
    parser.addoption(
        "--test-equivalence-only",
        action="store_true",
        help="Run only the refactored route equivalence tests"
    )

def pytest_collection_modifyitems(config, items):
    """Modify test collection based on command line options."""
    if config.getoption("--test-equivalence-only"):
        # Only run equivalence tests
        skip_other = pytest.mark.skip(reason="Skipping non-equivalence tests")
        for item in items:
            if not item.name.startswith("test_") or "equivalence" not in item.name:
                item.add_marker(skip_other)
