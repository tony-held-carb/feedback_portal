import os
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
import xlwings as xw

FILE_A = "diagnostics/dairy_digester_operator_feedback_v006_for_review_local.xlsx"
FILE_B = "diagnostics/dairy_digester_operator_feedback_v006_for_review_sharepoint.xlsx"


def print_header(title):
    print("\n" + "=" * 60)
    print(f"{title}")
    print("=" * 60)

def compare_file_metadata(path1, path2):
    print_header("File Metadata Comparison")
    stat1 = os.stat(path1)
    stat2 = os.stat(path2)
    print(f"{path1} size: {stat1.st_size} bytes, mtime: {datetime.fromtimestamp(stat1.st_mtime)}")
    print(f"{path2} size: {stat2.st_size} bytes, mtime: {datetime.fromtimestamp(stat2.st_mtime)}")
    if stat1.st_size != stat2.st_size:
        print("[!] File sizes differ!")
    if stat1.st_mtime != stat2.st_mtime:
        print("[!] Modification times differ!")
    else:
        print("File sizes and modification times are identical.")

    # Try to read workbook properties
    try:
        wb1 = openpyxl.load_workbook(path1, read_only=True, data_only=True)
        wb2 = openpyxl.load_workbook(path2, read_only=True, data_only=True)
        props1 = wb1.properties
        props2 = wb2.properties
        for attr in dir(props1):
            if not attr.startswith("_") and hasattr(props2, attr):
                v1 = getattr(props1, attr)
                v2 = getattr(props2, attr)
                if v1 != v2:
                    print(f"[!] Workbook property '{attr}' differs: {v1} vs {v2}")
        wb1.close()
        wb2.close()
    except Exception as e:
        print(f"[!] Could not read workbook properties: {e}")

def binary_diff(path1, path2):
    print_header("Binary Diff (byte-by-byte)")
    with open(path1, "rb") as f1, open(path2, "rb") as f2:
        chunk = 4096
        offset = 0
        diff_found = False
        while True:
            b1 = f1.read(chunk)
            b2 = f2.read(chunk)
            if not b1 and not b2:
                break
            if b1 != b2:
                print(f"[!] Files differ at byte offset {offset}")
                diff_found = True
                break
            offset += chunk
        if not diff_found:
            print("Files are binary identical.")

def openpyxl_compare(path1, path2):
    print_header("openpyxl Comparison (values, formulas, basic formatting)")
    try:
        wb1 = openpyxl.load_workbook(path1, data_only=False)
        wb2 = openpyxl.load_workbook(path2, data_only=False)
    except Exception as e:
        print(f"[!] Could not open files with openpyxl: {e}")
        return
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    if sheets1 != sheets2:
        print(f"[!] Sheet names/order differ: {sheets1} vs {sheets2}")
    else:
        print("Sheet names/order are identical.")
    for sheet in sheets1:
        if sheet not in wb2:
            print(f"[!] Sheet '{sheet}' missing in file B")
            continue
        ws1 = wb1[sheet]
        ws2 = wb2[sheet]
        max_row = max(ws1.max_row, ws2.max_row)
        max_col = max(ws1.max_column, ws2.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                c1 = ws1.cell(row, col)
                c2 = ws2.cell(row, col)
                if c1.value != c2.value:
                    print(f"[!] {sheet} {c1.coordinate}: Value differs: {c1.value} vs {c2.value}")
                if c1.data_type != c2.data_type:
                    print(f"[!] {sheet} {c1.coordinate}: Data type differs: {c1.data_type} vs {c2.data_type}")
                if c1.has_style or c2.has_style:
                    if c1.font != c2.font:
                        print(f"[!] {sheet} {c1.coordinate}: Font differs.")
                    if c1.fill != c2.fill:
                        print(f"[!] {sheet} {c1.coordinate}: Fill differs.")
                    if c1.number_format != c2.number_format:
                        print(f"[!] {sheet} {c1.coordinate}: Number format differs: {c1.number_format} vs {c2.number_format}")
    wb1.close()
    wb2.close()
    print("openpyxl comparison complete.")

def xlwings_compare(path1, path2):
    print_header("xlwings Comparison (full Excel fidelity)")
    try:
        with xw.App(visible=False) as app:
            wb1 = app.books.open(str(path1))
            wb2 = app.books.open(str(path2))
            sheets1 = {ws.name for ws in wb1.sheets}
            sheets2 = {ws.name for ws in wb2.sheets}
            if sheets1 != sheets2:
                print(f"[!] Sheet names differ: {sheets1} vs {sheets2}")
            else:
                print("Sheet names are identical.")
            for sheet in sheets1 & sheets2:
                ws1 = wb1.sheets[sheet]
                ws2 = wb2.sheets[sheet]
                max_row = max(ws1.used_range.last_cell.row, ws2.used_range.last_cell.row)
                max_col = max(ws1.used_range.last_cell.column, ws2.used_range.last_cell.column)
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        c1 = ws1.range((row, col))
                        c2 = ws2.range((row, col))
                        if c1.value != c2.value:
                            print(f"[!] {sheet} {c1.get_address(False, False)}: Value differs: {c1.value} vs {c2.value}")
                        if c1.formula != c2.formula:
                            print(f"[!] {sheet} {c1.get_address(False, False)}: Formula differs: {c1.formula} vs {c2.formula}")
                        # Formatting
                        if c1.api.Font.Name != c2.api.Font.Name:
                            print(f"[!] {sheet} {c1.get_address(False, False)}: Font name differs: {c1.api.Font.Name} vs {c2.api.Font.Name}")
                        if c1.api.Interior.Color != c2.api.Interior.Color:
                            print(f"[!] {sheet} {c1.get_address(False, False)}: Fill color differs: {c1.api.Interior.Color} vs {c2.api.Interior.Color}")
            wb1.close()
            wb2.close()
    except Exception as e:
        print(f"[!] xlwings comparison failed: {e}")
    print("xlwings comparison complete.")

def main():
    path1 = FILE_A
    path2 = FILE_B
    if not (os.path.exists(path1) and os.path.exists(path2)):
        print(f"ERROR: One or both files not found: {path1}, {path2}")
        sys.exit(1)
    compare_file_metadata(path1, path2)
    binary_diff(path1, path2)
    openpyxl_compare(path1, path2)
    xlwings_compare(path1, path2)

if __name__ == "__main__":
    main() 