"""
Automated Test Data Generation for Feedback Form Excel Ingestion

This script generates Excel files for all supported feedback form schemas (dairy, landfill, oil & gas, energy, generic),
covering a variety of scenarios:
- Valid data
- Missing required fields
- Invalid data types
- Extra/unexpected fields
- Large datasets
- Conditional logic triggers (contingent/cross-field)

Generated files are saved to: feedback_forms/testing_versions/generated/
A manifest JSON is created describing each file and its intended scenario.

Usage:
  python scripts/generate_test_excel_files.py

Requirements:
  - openpyxl

"""
import os
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import string

# Output directory
OUTPUT_DIR = Path("feedback_forms/testing_versions/generated")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Manifest to describe generated files
manifest = {}

# Supported schemas and their minimal required fields for demonstration
SCHEMAS = {
    "dairy": {
        "fields": [
            ("Facility Name", "Test Dairy Facility"),
            ("id_plume", 123),
            ("lat_carb", 35.3211),
            ("long_carb", -119.5808),
            ("Feedback Form", "Dairy Feedback"),
        ]
    },
    "landfill": {
        "fields": [
            ("Facility Name", "Test Landfill Facility"),
            ("id_plume", 456),
            ("lat_carb", 36.1234),
            ("long_carb", -120.1234),
            ("Feedback Form", "Landfill Feedback"),
        ]
    },
    "oil_and_gas": {
        "fields": [
            ("Facility Name", "Test O&G Facility"),
            ("id_plume", 789),
            ("lat_carb", 34.5678),
            ("long_carb", -118.5678),
            ("Feedback Form", "Oil & Gas Feedback"),
        ]
    },
    "energy": {
        "fields": [
            ("Facility Name", "Test Energy Facility"),
            ("id_plume", 321),
            ("lat_carb", 37.0000),
            ("long_carb", -121.0000),
            ("Feedback Form", "Energy Feedback"),
        ]
    },
    "generic": {
        "fields": [
            ("Facility Name", "Test Generic Facility"),
            ("id_plume", 654),
            ("lat_carb", 38.1111),
            ("long_carb", -122.1111),
            ("Feedback Form", "Generic Feedback"),
        ]
    },
}

# Helper to write a basic Feedback Form worksheet
def write_feedback_form(ws, fields):
    row = 15
    for key, value in fields:
        ws[f"B{row}"] = key
        ws[f"C{row}"] = value
        row += 1

# Helper to add extra/unexpected fields
def add_extra_fields(ws, start_row=30, count=5):
    for i in range(count):
        ws[f"B{start_row + i}"] = f"ExtraField{i+1}"
        ws[f"C{start_row + i}"] = f"ExtraValue{i+1}"

# Helper to add large number of rows
def add_large_data(ws, start_row=40, count=100):
    for i in range(count):
        ws[f"B{start_row + i}"] = f"Row{i+1}"
        ws[f"C{start_row + i}"] = random.randint(1, 10000)

# Helper to add conditional logic triggers
def add_conditional_logic(ws, schema):
    if schema == "landfill":
        ws["B25"] = "emission_identified_flag_fk"
        ws["C25"] = "No leak was detected"
        ws["B26"] = "emission_type_fk"
        ws["C26"] = "Type A"  # Should trigger cross-field logic
    elif schema == "oil_and_gas":
        ws["B25"] = "venting_exclusion"
        ws["C25"] = "Yes"
        ws["B26"] = "ogi_performed"
        ws["C26"] = "No"

# Helper to add invalid types
def add_invalid_types(ws):
    ws["C15"] = "NotANumber"
    ws["C16"] = "NotALatitude"
    ws["C17"] = "NotALongitude"

# Helper to remove required fields
def remove_required_fields(ws):
    ws["C15"] = ""
    ws["C16"] = ""

# Main generator loop
for schema, info in SCHEMAS.items():
    # 1. Valid file
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Form"
    write_feedback_form(ws, info["fields"])
    valid_path = OUTPUT_DIR / f"{schema}_valid.xlsx"
    wb.save(valid_path)
    manifest[str(valid_path)] = {"schema": schema, "scenario": "valid"}

    # 2. Missing required fields
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Form"
    write_feedback_form(ws, info["fields"])
    remove_required_fields(ws)
    missing_path = OUTPUT_DIR / f"{schema}_missing_required.xlsx"
    wb.save(missing_path)
    manifest[str(missing_path)] = {"schema": schema, "scenario": "missing_required"}

    # 3. Invalid data types
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Form"
    write_feedback_form(ws, info["fields"])
    add_invalid_types(ws)
    invalid_path = OUTPUT_DIR / f"{schema}_invalid_types.xlsx"
    wb.save(invalid_path)
    manifest[str(invalid_path)] = {"schema": schema, "scenario": "invalid_types"}

    # 4. Extra/unexpected fields
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Form"
    write_feedback_form(ws, info["fields"])
    add_extra_fields(ws)
    extra_path = OUTPUT_DIR / f"{schema}_extra_fields.xlsx"
    wb.save(extra_path)
    manifest[str(extra_path)] = {"schema": schema, "scenario": "extra_fields"}

    # 5. Large dataset
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Form"
    write_feedback_form(ws, info["fields"])
    add_large_data(ws)
    large_path = OUTPUT_DIR / f"{schema}_large.xlsx"
    wb.save(large_path)
    manifest[str(large_path)] = {"schema": schema, "scenario": "large"}

    # 6. Conditional logic triggers
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Form"
    write_feedback_form(ws, info["fields"])
    add_conditional_logic(ws, schema)
    cond_path = OUTPUT_DIR / f"{schema}_conditional_logic.xlsx"
    wb.save(cond_path)
    manifest[str(cond_path)] = {"schema": schema, "scenario": "conditional_logic"}

# Write manifest
manifest_path = OUTPUT_DIR / "manifest.json"
with open(manifest_path, "w") as f:
    json.dump(manifest, f, indent=2)

print(f"Generated {len(manifest)} test Excel files in {OUTPUT_DIR}")
print(f"Manifest written to {manifest_path}")

import os
from openpyxl import Workbook
import random
import string

EDGE_CASE_DIR = "feedback_forms/testing_versions/edge_cases"
os.makedirs(EDGE_CASE_DIR, exist_ok=True)

# 1. Corrupted Excel file (actually a text file with .xlsx extension)
with open(os.path.join(EDGE_CASE_DIR, "corrupted_file.xlsx"), "w") as f:
    f.write("This is not a real Excel file.")

# 2. Missing columns (missing required fields)
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
ws.append(["Row", "Field", "Value"])
ws.append([15, "SomeField", "SomeValue"])  # Missing required fields
wb.save(os.path.join(EDGE_CASE_DIR, "missing_columns.xlsx"))

# 3. Extra columns
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
ws.append(["Row", "Field", "Value", "Extra1", "Extra2"])
ws.append([15, "id_incidence", 123, "foo", "bar"])
ws.append([16, "SomeField", "SomeValue", "baz", "qux"])
wb.save(os.path.join(EDGE_CASE_DIR, "extra_columns.xlsx"))

# 4. Duplicate fields
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
ws.append(["Row", "Field", "Value"])
ws.append([15, "id_incidence", 123])
ws.append([16, "id_incidence", 456])  # Duplicate field
wb.save(os.path.join(EDGE_CASE_DIR, "duplicate_fields.xlsx"))

# 5. Large file (1000 rows)
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
ws.append(["Row", "Field", "Value"])
for i in range(15, 1015):
    ws.append([i, f"Field{i}", f"Value{i}"])
wb.save(os.path.join(EDGE_CASE_DIR, "large_file.xlsx"))

# 6. Unicode/special characters
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
ws.append(["Row", "Field", "Value"])
ws.append([15, "id_incidence", 123])
ws.append([16, "EmojiField", "😀🚀✨"])
ws.append([17, "RTLField", "مرحبا"])
ws.append([18, "CJKField", "漢字"])
wb.save(os.path.join(EDGE_CASE_DIR, "unicode_fields.xlsx"))

# 7. Mixed data types
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
ws.append(["Row", "Field", "Value"])
ws.append([15, "id_incidence", 123])
ws.append([16, "NumericField", "not_a_number"])  # Should be numeric
ws.append([17, "DateField", "not_a_date"])      # Should be date
wb.save(os.path.join(EDGE_CASE_DIR, "mixed_types.xlsx"))

# 8. Headers only
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
ws.append(["Row", "Field", "Value"])
wb.save(os.path.join(EDGE_CASE_DIR, "headers_only.xlsx"))

# 9. Data only (no headers)
wb = Workbook()
ws = wb.active
ws.title = "Feedback Form"
for i in range(15, 18):
    ws.append([i, f"Field{i}", f"Value{i}"])
wb.save(os.path.join(EDGE_CASE_DIR, "data_only.xlsx"))

print("Edge case Excel files generated in:", EDGE_CASE_DIR) 